# -*- coding: utf-8 -*-
import io
import os
import shutil
import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List

import models
import schemas
from database import get_db
from services.excel_classifier import get_group_code
from services.form_validator import validate_form
from services.file_merger import merge_file

router = APIRouter(
    prefix="/api/files",
    tags=["files"],
)

UPLOAD_DIR = "uploads"

@router.post("/upload", response_model=schemas.FileQueue)
async def upload_file(
    quarter_id: str = Form(...),
    sheet_name: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    if not file.filename.lower().endswith(('.xls', '.xlsx', '.xlsb', '.xlsm')):
        raise HTTPException(status_code=400, detail="Only Excel files are allowed.")
        
    quarter_dir = os.path.join(UPLOAD_DIR, quarter_id)
    os.makedirs(quarter_dir, exist_ok=True)
    file_path = os.path.join(quarter_dir, file.filename)
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    group_code = get_group_code(file.filename, sheet_name)
    if not group_code:
        raise HTTPException(status_code=400, detail="Cannot determine Group Code from file name and sheet name. Ensure keywords exist.")

    term = "N/A"
    if group_code.endswith("_LT"):
        term = "Long Term"
    elif group_code.endswith("_ST"):
        term = "Short Term"
        
    existing_file = db.query(models.FileQueue).filter(
        models.FileQueue.quarter_id == quarter_id,
        models.FileQueue.file_name == file.filename,
        models.FileQueue.sheet_name == sheet_name
    ).first()
    
    if existing_file:
        existing_file.file_path = file_path
        existing_file.group_code = group_code
        existing_file.term = term
        existing_file.status = "Pending"
        db.commit()
        db.refresh(existing_file)
        return existing_file
    else:
        new_file = models.FileQueue(
            quarter_id=quarter_id,
            file_name=file.filename,
            sheet_name=sheet_name,
            file_path=file_path,
            group_code=group_code,
            term=term,
            status="Pending"
        )
        db.add(new_file)
        db.commit()
        db.refresh(new_file)
        return new_file

@router.get("/", response_model=List[schemas.FileQueue])
def list_files(quarter_id: str, db: Session = Depends(get_db)):
    files = db.query(models.FileQueue).filter(models.FileQueue.quarter_id == quarter_id).all()
    return files

@router.post("/{file_id}/check")
def check_file(file_id: int, db: Session = Depends(get_db)):
    file_record = db.query(models.FileQueue).filter(models.FileQueue.id == file_id).first()
    if not file_record:
        raise HTTPException(status_code=404, detail="File not found in queue.")

    res = validate_form(file_record.file_path, file_record.sheet_name, file_record.group_code)

    if res["ok"] and not res.get("errors"):
        # Only mark Validated when no fatal errors; warnings are allowed
        file_record.status = "Validated"
        db.commit()

    # Drop non-serializable 'df' and sanitize numpy/NaN/inf before returning JSON
    import math, json
    from fastapi.responses import JSONResponse

    def _make_safe(obj):
        """Recursively convert numpy scalars, NaN, inf to JSON-safe Python types."""
        if obj is None:
            return None
        # numpy scalar types
        type_name = type(obj).__name__
        if type_name in ("int64", "int32", "int16", "int8", "uint64", "uint32", "uint16", "uint8"):
            return int(obj)
        if type_name in ("float64", "float32", "float16"):
            if math.isnan(float(obj)) or math.isinf(float(obj)):
                return None
            return float(obj)
        if isinstance(obj, float):
            if math.isnan(obj) or math.isinf(obj):
                return None
            return obj
        if isinstance(obj, dict):
            return {str(k): _make_safe(v) for k, v in obj.items()}
        if isinstance(obj, (list, tuple)):
            return [_make_safe(i) for i in obj]
        return obj

    safe_res = _make_safe({k: v for k, v in res.items() if k != "df"})
    return JSONResponse(content=safe_res)

@router.get("/{file_id}/check-report")
def download_check_report(file_id: int, db: Session = Depends(get_db)):
    """
    Validation summary report (stats + warnings). See /download-checked for full data.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    file_record = db.query(models.FileQueue).filter(models.FileQueue.id == file_id).first()
    if not file_record:
        raise HTTPException(status_code=404, detail="File not found.")

    res = validate_form(file_record.file_path, file_record.sheet_name, file_record.group_code)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validation Report"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    ok_font    = Font(color="059669", bold=True)
    err_font   = Font(color="DC2626", bold=True)
    warn_fill  = PatternFill("solid", fgColor="FEF9C3")
    thin = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:C1")
    ws["A1"] = "Validation Report - " + file_record.file_name
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    meta = [
        ("File", file_record.file_name),
        ("Sheet", file_record.sheet_name),
        ("Group Code", file_record.group_code),
        ("Status", "PASSED" if res["ok"] else "FAILED"),
    ]
    for i, (k, v) in enumerate(meta, start=2):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    stat_start = len(meta) + 3
    ws.cell(row=stat_start, column=1, value="Metric").fill = header_fill
    ws.cell(row=stat_start, column=1).font = header_font
    ws.cell(row=stat_start, column=2, value="Value").fill = header_fill
    ws.cell(row=stat_start, column=2).font = header_font

    stats = res.get("stats", {})
    stat_rows = [
        ("Total rows",     stats.get("total_rows", "-")),
        ("Date errors",    stats.get("date_errors", 0)),
        ("Money errors",   stats.get("money_errors", 0)),
        ("Duplicate rows", stats.get("duplicate_rows", 0)),
    ]
    for j, (label, value) in enumerate(stat_rows, start=stat_start + 1):
        c_label = ws.cell(row=j, column=1, value=label)
        c_value = ws.cell(row=j, column=2, value=value)
        c_label.border = thin
        c_value.border = thin
        is_err = isinstance(value, int) and value > 0 and label != "Total rows"
        c_value.font = err_font if is_err else ok_font

    warnings_list = res.get("warnings", [])
    if warnings_list:
        warn_start = stat_start + len(stat_rows) + 2
        ws.cell(row=warn_start, column=1, value="Warnings").font = Font(bold=True)
        for k, w in enumerate(warnings_list, start=warn_start + 1):
            cell = ws.cell(row=k, column=1, value=w)
            cell.fill = warn_fill
            ws.merge_cells(start_row=k, start_column=1, end_row=k, end_column=3)

    errors_list = res.get("errors", [])
    if errors_list:
        err_start = stat_start + len(stat_rows) + len(warnings_list) + 3
        ws.cell(row=err_start, column=1, value="Errors").font = Font(bold=True, color="DC2626")
        for k, e in enumerate(errors_list, start=err_start + 1):
            cell = ws.cell(row=k, column=1, value=e)
            cell.font = Font(color="DC2626")
            ws.merge_cells(start_row=k, start_column=1, end_row=k, end_column=3)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    import unicodedata as _ud
    safe_name = file_record.file_name.replace(".xlsx", "").replace(".xls", "")
    filename_xlsx = f"ValidationReport_{safe_name}.xlsx"
    ascii_name = _ud.normalize("NFKD", filename_xlsx).encode("ascii", "ignore").decode("ascii")
    ascii_name = ascii_name.replace(" ", "_") or f"ValidationReport_{file_record.id}.xlsx"
    encoded_name = quote(filename_xlsx, safe="")

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{encoded_name}"},
    )


@router.get("/{file_id}/download-checked")
def download_checked_excel(file_id: int, db: Session = Depends(get_db)):
    """
    Mirrors R's download_checked_excel2 downloadHandler (lines 727-765):
    - Runs validation, gets annotated DataFrame with Check_Ngay_Hop_Le, Check_So_Tien, check_trung
    - Exports full data rows to Excel
    - Colors: Yellow = date error, Red = money error, Green = duplicate
    Filename: <group_code>_check_ngay_<date>.xlsx
    """
    import openpyxl
    from openpyxl.styles import PatternFill
    from datetime import date as _date_cls

    file_record = db.query(models.FileQueue).filter(models.FileQueue.id == file_id).first()
    if not file_record:
        raise HTTPException(status_code=404, detail="File not found.")

    res = validate_form(file_record.file_path, file_record.sheet_name, file_record.group_code)

    if not res.get("ok"):
        raise HTTPException(status_code=422, detail=res.get("errors", ["Validation failed"]))

    df_out = res.get("df")
    if df_out is None or (hasattr(df_out, "empty") and df_out.empty):
        raise HTTPException(status_code=422, detail="No data available to download.")

    # ── Build workbook mirroring R createWorkbook + addStyle ──────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Check"

    # Write header row
    headers = list(df_out.columns)
    for col_idx, col_name in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data rows
    for row_idx, (_, row) in enumerate(df_out.iterrows(), start=2):
        for col_idx, col_name in enumerate(headers, start=1):
            val = row[col_name]
            # Convert pandas NA/NaT to None for Excel
            if pd.isna(val) if not isinstance(val, str) else False:
                val = None
            ws.cell(row=row_idx, column=col_idx, value=val)

    # ── Apply color fills (mirrors R addStyle logic, lines 739-763) ──────
    # Yellow: date error  (#FFFF00)
    # Green:  duplicate   (#00FF00)
    # Red:    money error (#FF0000)
    # Priority (same as R): yellow first, then green, then red (red overwrites)
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    green_fill  = PatternFill("solid", fgColor="00FF00")
    red_fill    = PatternFill("solid", fgColor="FF0000")
    n_cols = len(headers)

    for row_idx, (_, row) in enumerate(df_out.iterrows(), start=2):
        check_ngay = str(row.get("Check_Ngay_Hop_Le", "Ngày hợp lệ"))
        check_tien = str(row.get("Check_So_Tien", "Số tiền hợp lệ"))
        check_trung = row.get("check_trung", 0)

        fill = None
        if check_ngay != "Ngày hợp lệ":
            fill = yellow_fill
        if str(check_trung) == "Trùng":
            fill = green_fill
        if check_tien != "Số tiền hợp lệ":
            fill = red_fill   # red has highest priority (overwrites)

        if fill is not None:
            for col_idx in range(1, n_cols + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    import unicodedata as _ud
    today_str = _date_cls.today().isoformat()
    group_code = file_record.group_code or "unknown"
    filename_xlsx = f"{group_code}_check_ngay_{today_str}.xlsx"
    ascii_name = _ud.normalize("NFKD", filename_xlsx).encode("ascii", "ignore").decode("ascii")
    ascii_name = ascii_name.replace(" ", "_") or f"checked_{file_record.id}.xlsx"
    encoded_name = quote(filename_xlsx, safe="")

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{encoded_name}"},
    )

@router.post("/{file_id}/merge")
def merge_file_endpoint(file_id: int, db: Session = Depends(get_db)):
    """
    Normalise a validated Excel file and save it as parquet in cur_data/.
    Also merges with the previous quarter's data if available.
    """
    file_record = db.query(models.FileQueue).filter(models.FileQueue.id == file_id).first()
    if not file_record:
        raise HTTPException(status_code=404, detail="File not found in queue.")

    if file_record.status not in ("Validated", "Pending"):
        raise HTTPException(status_code=400, detail="File must be Validated before merging.")

    # Get calculation date from parameters (for end-date filtering)
    from datetime import date as _date
    param = db.query(models.AppParameter).filter(
        models.AppParameter.quarter_id == file_record.quarter_id
    ).first()
    dpnv_date = None
    if param:
        try:
            dpnv_date = _date(param.year, param.month, param.day)
        except Exception:
            pass

    res = merge_file(
        file_path=file_record.file_path,
        sheet_name=file_record.sheet_name,
        group_code=file_record.group_code,
        quarter_id=file_record.quarter_id,
        dpnv_date=dpnv_date,
    )

    if res["ok"]:
        file_record.status = "Merged"
        db.commit()

    # Make response JSON-safe (strip non-serialisable values)
    safe_res = {k: v for k, v in res.items() if isinstance(v, (bool, int, str, list, dict, type(None)))}
    return safe_res


@router.get("/{file_id}/download-ketqua")
def download_ketqua(file_id: int, db: Session = Depends(get_db)):
    """
    Mirrors R's output$download_che downloadHandler (2.ghep_file.R line 736-748):
    Exports the audit/comparison report (Loai_dong + Cac_cot_thay_doi) as a styled Excel file.
    Colors: green=Mới hoàn toàn, yellow=Thay đổi, orange=Trùng, red=Bị bỏ.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from datetime import date as _date_cls
    from services.file_merger import CUR_DATA_ROOT

    file_record = db.query(models.FileQueue).filter(models.FileQueue.id == file_id).first()
    if not file_record:
        raise HTTPException(status_code=404, detail="File not found.")

    # Build expected ketqua parquet path
    ketqua_path = os.path.join(
        CUR_DATA_ROOT,
        file_record.quarter_id,
        f"{file_record.group_code}_ketqua.parquet",
    )
    if not os.path.exists(ketqua_path):
        raise HTTPException(
            status_code=404,
            detail="Audit report not found. Please run the Merge step first."
        )

    df_kq = pd.read_parquet(ketqua_path)
    if df_kq.empty:
        raise HTTPException(status_code=422, detail="Audit report is empty — no changes detected between quarters.")

    # ── Build Excel workbook ──────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kiem_tra_thay_doi"

    # Header row
    headers = list(df_kq.columns)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    for ci, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.fill = header_fill
        cell.font = header_font

    # Color map matching R's colour semantics (adapted to audit categories)
    fill_map = {
        "Mới hoàn toàn": PatternFill("solid", fgColor="C6EFCE"),   # green
        "Thay đổi":       PatternFill("solid", fgColor="FFFF00"),   # yellow
        "Thêm thông tin":   PatternFill("solid", fgColor="FFEB9C"),   # light orange
        "Trùng":           PatternFill("solid", fgColor="FFC7CE"),   # light red/pink
        "Bị bỏ":          PatternFill("solid", fgColor="FF0000"),   # red
    }

    # Loai_dong column index (1-based)
    loai_dong_col = headers.index("Loai_dong") if "Loai_dong" in headers else None

    # Write data rows
    for ri, (_, row) in enumerate(df_kq.iterrows(), start=2):
        loai = str(row.get("Loai_dong", "")) if loai_dong_col is not None else ""
        fill = fill_map.get(loai)

        for ci, col_name in enumerate(headers, start=1):
            val = row[col_name]
            if pd.isna(val) if not isinstance(val, str) else (val in ("None", "nan", "<NA>")):
                val = None
            cell = ws.cell(row=ri, column=ci, value=val)
            if fill:
                cell.fill = fill

    # Auto-size key columns
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    import unicodedata as _ud
    today_str = _date_cls.today().isoformat()
    gc = file_record.group_code or "unknown"
    filename_xlsx = f"{gc}_check_{today_str}.xlsx"
    ascii_name = _ud.normalize("NFKD", filename_xlsx).encode("ascii", "ignore").decode("ascii")
    ascii_name = ascii_name.replace(" ", "_") or f"ketqua_{file_record.id}.xlsx"
    encoded_name = quote(filename_xlsx, safe="")

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{ascii_name}"; filename*=UTF-8\'\'{encoded_name}'},
    )

@router.delete("/clear")
def clear_files(quarter_id: str, db: Session = Depends(get_db)):
    db.query(models.FileQueue).filter(models.FileQueue.quarter_id == quarter_id).delete()
    db.commit()
    
    quarter_dir = os.path.join(UPLOAD_DIR, quarter_id)
    if os.path.exists(quarter_dir):
        shutil.rmtree(quarter_dir, ignore_errors=True)
        
    return {"ok": True, "message": "All files cleared."}
