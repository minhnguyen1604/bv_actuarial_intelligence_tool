import os
import re
import datetime
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
import io
from sqlalchemy.orm import Session

from database import get_db
import models
from services.upr_calculator import load_ty_gia

router = APIRouter(prefix="/api/premium-analysis", tags=["Premium Analysis"])

DATA_ROOT = r"d:\bv_intelligence_tool\python_code\backend\data"

# Helper list of LOBs
ST_LOBS = ['Eng_ST', 'Fire_ST', 'Marine_ST', 'Misc_ST', 'PA_ST', 'Travel_BHTT', 'Travel_CTTV', 'XCG_ST']
LT_LOBS = ['Eng_LT', 'Fire_LT', 'Marine_LT', 'Misc_LT', 'PA_LT', 'PA_TTTBVV', 'XCG_LT', 'XCG_CWVN_LT']

def get_quarters_for_date(analysis_date: datetime.date) -> tuple[str, str]:
    """Given a date, return the current quarter (e.g. Q1/2026) and previous quarter (e.g. Q4/2025)."""
    month = analysis_date.month
    year = analysis_date.year
    
    if 1 <= month <= 3:
        curr_q = f"Q1/{year}"
        prev_q = f"Q4/{year - 1}"
    elif 4 <= month <= 6:
        curr_q = f"Q2/{year}"
        prev_q = f"Q1/{year}"
    elif 7 <= month <= 9:
        curr_q = f"Q3/{year}"
        prev_q = f"Q2/{year}"
    else:
        curr_q = f"Q4/{year}"
        prev_q = f"Q3/{year}"
        
    return curr_q, prev_q

def get_quarter_id_from_str(q_str: str) -> str:
    """Convert Q1/2026 to Q1_2026."""
    return q_str.replace("/", "_")

def compute_lob_rows(group_code: str, quarter_id: str, dpnv_date: datetime.date, ty_gia: pd.DataFrame) -> pd.DataFrame:
    """
    Load merged parquet file and calculate row-level premium and UPR values.
    Returns a DataFrame containing Policy_Number, Quarter, Premium, UPR.
    """
    dot_qid = quarter_id.replace("_", ".")
    parquet_path = os.path.join(DATA_ROOT, dot_qid, f"{group_code}.parquet")
    if not os.path.exists(parquet_path):
        alt_path = os.path.join(DATA_ROOT, quarter_id, f"{group_code}.parquet")
        if os.path.exists(alt_path):
            parquet_path = alt_path
        else:
            return pd.DataFrame(columns=["Policy_Number", "Quarter", "Premium", "UPR", "Ky_phi"])

    df = pd.read_parquet(parquet_path)
    if df.empty:
        return pd.DataFrame(columns=["Policy_Number", "Quarter", "Premium", "UPR", "Ky_phi"])

    # Determine LOB characteristics
    group_lower = group_code.lower()
    is_lt = group_lower.endswith("_lt") or group_lower == "xcg_cwvn_lt" or group_lower == "pa_lt"

    # 1. Exchange rates mapping
    ty_gia_dict_usd = {k: v for k, v in zip(ty_gia["Thoi_gian"].astype(str), ty_gia["USD"]) if pd.notna(v)}
    ty_gia_dict_eur = {k: v for k, v in zip(ty_gia["Thoi_gian"].astype(str), ty_gia["EUR"]) if pd.notna(v)}
    fallback_usd = ty_gia["USD"].dropna().iloc[-1] if not ty_gia["USD"].dropna().empty else 1.0
    fallback_eur = ty_gia["EUR"].dropna().iloc[-1] if not ty_gia["EUR"].dropna().empty else 1.0


    # 2. Get installments (Ky_phi)
    ky_available = []
    for col in df.columns:
        m = re.match(r"^Ky_phi_(\d+)_", col)
        if m:
            ky_available.append(int(m.group(1)))
    ky_available = sorted(list(set(ky_available)))
    if not ky_available:
        return pd.DataFrame(columns=["Policy_Number", "Quarter", "Premium", "UPR", "Ky_phi"])

    dpnv = pd.to_datetime(dpnv_date)
    detail_rows = []

    # 3. Calculate based LOB type
    for ky in ky_available:
        # Check standard columns
        vnd_col = f"Ky_phi_{ky}_So_tien_VND"
        usd_col = f"Ky_phi_{ky}_So_tien_USD"
        eur_col = f"Ky_phi_{ky}_So_tien_EUR"
        thang_col = f"Ky_phi_{ky}_Ghi_Thang"
        nam_col = f"Ky_phi_{ky}_Ghi_Nam"

        if vnd_col not in df.columns or thang_col not in df.columns:
            continue

        thang = pd.to_numeric(df[thang_col], errors='coerce').fillna(0).astype(int)
        nam = pd.to_numeric(df[nam_col], errors='coerce').fillna(0).astype(int)
        
        quy = np.where(thang == 0, 0, (thang - 1) // 3 + 1)
        quy_nam = "Q" + pd.Series(quy).astype(str) + "/" + pd.Series(nam).astype(str)
        quy_nam = np.where((quy == 0) | (nam == 0), "", quy_nam)
        
        rate_usd = pd.Series(quy_nam).map(ty_gia_dict_usd).fillna(fallback_usd).values
        rate_eur = pd.Series(quy_nam).map(ty_gia_dict_eur).fillna(fallback_eur).values
        
        vnd = pd.to_numeric(df[vnd_col].astype(str).str.replace(",", "", regex=False), errors='coerce').fillna(0.0).values
        usd = pd.to_numeric(df[usd_col].astype(str).str.replace(",", "", regex=False), errors='coerce').fillna(0.0).values if usd_col in df.columns else np.zeros(len(df))
        eur = pd.to_numeric(df[eur_col].astype(str).str.replace(",", "", regex=False), errors='coerce').fillna(0.0).values if eur_col in df.columns else np.zeros(len(df))
        
        # Premium/Fee
        phi_goc = vnd + usd * rate_usd + eur * rate_eur

        # Get retention rate tile
        ret_col = "Ty_le_giu_lai_cua_BHBV" if "Ty_le_giu_lai_cua_BHBV" in df.columns else \
                  ("Ty_le_giu_lai_cua_BHBV_checked" if "Ty_le_giu_lai_cua_BHBV_checked" in df.columns else None)
        if ret_col:
            hi = pd.to_numeric(df[ret_col].astype(str).str.replace(",", "", regex=False), errors='coerce').fillna(0.0).values
            tile = np.where((hi == 0.0) | df[ret_col].isna(), 1.0, np.where(hi > 1.0, hi / 100.0, hi))
        else:
            tile = np.ones(len(df))

        phi_giulai = phi_goc * tile
        phi_tai = phi_goc - phi_giulai

        # Helper to parse dates
        def parse_date_series(d_col, m_col, y_col):
            if d_col not in df.columns:
                return pd.Series([pd.NaT]*len(df))
            d_val = pd.to_numeric(df[d_col], errors='coerce')
            m_val = pd.to_numeric(df[m_col], errors='coerce')
            y_val = pd.to_numeric(df[y_col], errors='coerce')
            d_str = y_val.fillna(2000).astype(int).astype(str).str.zfill(4) + '-' + \
                    m_val.fillna(1).astype(int).astype(str).str.zfill(2) + '-' + \
                    d_val.fillna(1).astype(int).astype(str).str.zfill(2)
            parsed = pd.to_datetime(d_str, format='%Y-%m-%d', errors='coerce')
            parsed = parsed.where(d_val.notna() & m_val.notna() & y_val.notna(), pd.NaT)
            return parsed

        if is_lt:
            # Long-Term logic
            tu_date = parse_date_series("Thoi_han_bao_hiem_Tu_Ngay", "Thoi_han_bao_hiem_Tu_Thang", "Thoi_han_bao_hiem_Tu_Nam")
            den_date = parse_date_series("Thoi_han_bao_hiem_Den_Ngay", "Thoi_han_bao_hiem_Den_Thang", "Thoi_han_bao_hiem_Den_Nam")
            
            tu_date_kp = parse_date_series(f"Ky_phi_{ky}_Tu_Ngay", f"Ky_phi_{ky}_Tu_Thang", f"Ky_phi_{ky}_Tu_Nam")
            den_date_kp = parse_date_series(f"Ky_phi_{ky}_Den_Ngay", f"Ky_phi_{ky}_Den_Thang", f"Ky_phi_{ky}_Den_Nam")
            ghi_date = parse_date_series(f"Ky_phi_{ky}_Ghi_Ngay", f"Ky_phi_{ky}_Ghi_Thang", f"Ky_phi_{ky}_Ghi_Nam")
            
            diff_dpnv_den = (dpnv - den_date_kp).dt.days
            c1 = np.where(diff_dpnv_den.isna(), 0, np.where(diff_dpnv_den >= 0, 0, 1))
            
            diff_den_tu = (den_date - tu_date).dt.days
            c2 = np.where(diff_den_tu.isna() | (diff_den_tu <= 365), 0, 1)
            
            c3 = np.where(df[vnd_col].isna(), 0, 1)
            
            diff_kp = (den_date_kp - tu_date_kp).dt.days
            c4 = np.where(diff_kp.isna(), 0, np.where(diff_kp > 0, 1, 0))
            
            c5 = np.where(df[thang_col].isna() & df[nam_col].isna(), 0, 1)
            
            diff_dpnv_tu = (dpnv - tu_date).dt.days
            c6 = np.where(diff_dpnv_tu.isna(), 0, np.where(diff_dpnv_tu >= 0, 1, 0))
            
            diff_ghi_dpnv = (ghi_date - dpnv).dt.days
            c7 = np.where(diff_ghi_dpnv.isna(), 0, np.where(diff_ghi_dpnv > 0, 0, 1))
            
            tonghop = np.where((c1 == 0) | (c2 == 0) | (c3 == 0) | (c4 == 0) | (c5 == 0) | (c6 == 0) | (c7 == 0), 0, 1)
            
            tu_nam_kp = pd.to_numeric(df[f"Ky_phi_{ky}_Tu_Nam"], errors='coerce').fillna(0).astype(int)
            tu_thang_kp = pd.to_numeric(df[f"Ky_phi_{ky}_Tu_Thang"], errors='coerce').fillna(0).astype(int)
            thu_tu = np.where(
                tonghop == 0,
                0,
                (dpnv_date.year - tu_nam_kp) * 4 + (dpnv_date.month - 1) // 3 + 1 - ((tu_thang_kp - 1) // 3 + 1) + 1
            )
            
            diff_kp_1 = (den_date_kp - tu_date_kp).dt.days + 1
            mau_so = np.where(diff_kp_1.isna(), 0.0, (diff_kp_1 / 365.0) * 8.0)
            
            huong_cu = np.where(tonghop == 0, mau_so, np.where(thu_tu <= 0, 0.0, thu_tu * 2.0 - 1.0))
            chua_huong = mau_so - huong_cu
            
            diff_den_dpnv = (den_date_kp - dpnv).dt.days
            chua_huong_dc = np.where(
                chua_huong >= 0.0,
                chua_huong,
                np.where(diff_den_dpnv.isna(), 0.0, (diff_den_dpnv / 365.0) * 8.0)
            )
            
            upr_ratio = np.where(mau_so == 0.0, 0.0, chua_huong_dc / np.where(mau_so == 0.0, 1.0, mau_so))
            upr_val = phi_goc * upr_ratio
            
            for idx in range(len(df)):
                if tonghop[idx] == 1:
                    detail_rows.append({
                        "Policy_Number": str(df["So_don_Ma_hop_dong_Ma_SDBS"].iloc[idx]),
                        "Quarter": quy_nam[idx],
                        "Premium": phi_goc[idx],
                        "UPR": upr_val[idx],
                        "Ky_phi": ky
                    })
        else:
            # Short-Term / Standard / Vietjet / TTTBVV logic
            tu_date = parse_date_series("Thoi_han_bao_hiem_Tu_Ngay", "Thoi_han_bao_hiem_Tu_Thang", "Thoi_han_bao_hiem_Tu_Nam")
            den_date = parse_date_series("Thoi_han_bao_hiem_Den_Ngay", "Thoi_han_bao_hiem_Den_Thang", "Thoi_han_bao_hiem_Den_Nam")
            
            diff_days = (den_date - tu_date).dt.days
            dem_ngay = np.where(diff_days.isna(), 0, np.where(diff_days < 365, 1, 0))
            
            diff_dpnv_den = (dpnv - den_date).dt.days
            hethieuluc = np.where(diff_dpnv_den.isna(), 0, np.where(diff_dpnv_den >= 0, 1, 0))
            
            upr_val = np.where((dem_ngay == 1) & (hethieuluc == 1), phi_goc, 0.0)
            
            for idx in range(len(df)):
                q_val = quy_nam[idx]
                if q_val != "":
                    detail_rows.append({
                        "Policy_Number": str(df["So_don_Ma_hop_dong_Ma_SDBS"].iloc[idx]),
                        "Quarter": q_val,
                        "Premium": phi_goc[idx],
                        "UPR": upr_val[idx],
                        "Ky_phi": ky
                    })

    return pd.DataFrame(detail_rows)

def query_db_lob_metrics(db_filename: str, lob: str, quarter_name: str) -> dict:
    metrics = {
        "gross_premium": 0.0,
        "net_premium": 0.0,
        "rein_premium": 0.0,
        "net_upr": 0.0,
        "rein_upr": 0.0
    }
    
    db_path = os.path.join(DATA_ROOT, db_filename)
    if not os.path.exists(db_path):
        return metrics
        
    conn = sqlite3.connect(db_path)
    try:
        cursor = conn.cursor()
        
        # Determine table and query type
        is_tttbvv = "tttbvv" in lob.lower()
        is_lt = lob.lower().endswith("_lt") or lob.lower() == "xcg_cwvn_lt"
        
        if is_tttbvv:
            # Query PA_TTTBVV_Summary table filtering by Quy = quarter_name
            cursor.execute(
                "SELECT SUM(Phi_bao_hiem_goc), SUM(Phi_bao_hiem_giu_lai), "
                "SUM(Du_phong_bao_hiem_giu_lai), SUM(Du_phong_bao_hiem_tai) "
                "FROM PA_TTTBVV_Summary WHERE lob = ? AND Quy = ?",
                (lob, quarter_name)
            )
            row = cursor.fetchone()
            if row and row[0] is not None:
                goc = float(row[0] or 0.0)
                giulai = float(row[1] or 0.0)
                metrics["gross_premium"] = goc
                metrics["net_premium"] = giulai
                metrics["rein_premium"] = goc - giulai
                metrics["net_upr"] = float(row[2] or 0.0)
                metrics["rein_upr"] = float(row[3] or 0.0)
                
        elif is_lt:
            # Query Long_term table filtering by Quy = 'Số dùng để tính'
            cursor.execute(
                "SELECT SUM(Phi_bao_hiem_sau_dong), SUM(Phi_bao_hiem_giu_lai), SUM(Phi_tai_bao_hiem), "
                "SUM(Phi_bao_hiem_giu_lai_chua_huong), SUM(Phi_tai_bao_hiem_chua_huong) "
                "FROM Long_term WHERE lob = ? AND Quy = 'Số dùng để tính'",
                (lob,)
            )
            row = cursor.fetchone()
            if row and row[0] is not None:
                metrics["gross_premium"] = float(row[0] or 0.0)
                metrics["net_premium"] = float(row[1] or 0.0)
                metrics["rein_premium"] = float(row[2] or 0.0)
                metrics["net_upr"] = float(row[3] or 0.0)
                metrics["rein_upr"] = float(row[4] or 0.0)
                
        else:
            # Query Short_term table filtering by Quy = quarter_name
            cursor.execute(
                "SELECT SUM(Phi_bao_hiem_goc), SUM(Phi_bao_hiem_giu_lai), "
                "SUM(Giam_phi_bao_hiem_giu_lai), SUM(Giam_phi_bao_hiem_tai) "
                "FROM Short_term WHERE lob = ? AND Quy = ?",
                (lob, quarter_name)
            )
            row = cursor.fetchone()
            if row and row[0] is not None:
                goc = float(row[0] or 0.0)
                giulai = float(row[1] or 0.0)
                metrics["gross_premium"] = goc
                metrics["net_premium"] = giulai
                metrics["rein_premium"] = goc - giulai
                metrics["net_upr"] = float(row[2] or 0.0)
                metrics["rein_upr"] = float(row[3] or 0.0)
                
    except Exception as e:
        print(f"Error querying metrics for {lob} from {db_filename}: {e}")
    finally:
        conn.close()
        
    return metrics

import sqlite3

@router.get("/overview")
def get_premium_analysis_overview(day: int, month: int, year: int, db: Session = Depends(get_db)):
    """Retrieve Short-Term and Long-Term Premium and UPR overview table data from SQLite DB."""
    analysis_date = datetime.date(year, month, day)
    curr_q, prev_q = get_quarters_for_date(analysis_date)
    curr_qid = get_quarter_id_from_str(curr_q)
    prev_qid = get_quarter_id_from_str(prev_q)

    curr_db_file = f"{curr_qid.replace('_', '.')}.db"
    prev_db_file = f"{prev_qid.replace('_', '.')}.db"

    st_results = []
    lt_results = []

    # 1. Process Short-Term LOBs
    for lob in ST_LOBS:
        curr_metrics = query_db_lob_metrics(curr_db_file, lob, curr_q)
        prev_metrics = query_db_lob_metrics(prev_db_file, lob, prev_q)
        
        # Calculate % changes
        pct_gross = 0.0
        if prev_metrics["gross_premium"] > 0:
            pct_gross = ((curr_metrics["gross_premium"] - prev_metrics["gross_premium"]) / prev_metrics["gross_premium"]) * 100.0
            
        pct_net = 0.0
        if prev_metrics["net_premium"] > 0:
            pct_net = ((curr_metrics["net_premium"] - prev_metrics["net_premium"]) / prev_metrics["net_premium"]) * 100.0
            
        pct_rein = 0.0
        if prev_metrics["rein_premium"] > 0:
            pct_rein = ((curr_metrics["rein_premium"] - prev_metrics["rein_premium"]) / prev_metrics["rein_premium"]) * 100.0
            
        pct_net_upr = 0.0
        if prev_metrics["net_upr"] > 0:
            pct_net_upr = ((curr_metrics["net_upr"] - prev_metrics["net_upr"]) / prev_metrics["net_upr"]) * 100.0
            
        pct_rein_upr = 0.0
        if prev_metrics["rein_upr"] > 0:
            pct_rein_upr = ((curr_metrics["rein_upr"] - prev_metrics["rein_upr"]) / prev_metrics["rein_upr"]) * 100.0

        st_results.append({
            "lob": lob,
            # Old backward-compatibility keys
            "prev_premium": float(prev_metrics["gross_premium"]),
            "curr_premium": float(curr_metrics["gross_premium"]),
            "pct_change": float(pct_gross),
            "prev_upr": float(prev_metrics["net_upr"]),
            "curr_upr": float(curr_metrics["net_upr"]),
            # New detailed keys
            "prev_gross_premium": float(prev_metrics["gross_premium"]),
            "curr_gross_premium": float(curr_metrics["gross_premium"]),
            "pct_gross_premium": float(pct_gross),
            "prev_net_premium": float(prev_metrics["net_premium"]),
            "curr_net_premium": float(curr_metrics["net_premium"]),
            "pct_net_premium": float(pct_net),
            "prev_rein_premium": float(prev_metrics["rein_premium"]),
            "curr_rein_premium": float(curr_metrics["rein_premium"]),
            "pct_rein_premium": float(pct_rein),
            "prev_net_upr": float(prev_metrics["net_upr"]),
            "curr_net_upr": float(curr_metrics["net_upr"]),
            "pct_net_upr": float(pct_net_upr),
            "prev_rein_upr": float(prev_metrics["rein_upr"]),
            "curr_rein_upr": float(curr_metrics["rein_upr"]),
            "pct_rein_upr": float(pct_rein_upr)
        })

    # 2. Process Long-Term LOBs
    for lob in LT_LOBS:
        curr_metrics = query_db_lob_metrics(curr_db_file, lob, curr_q)
        prev_metrics = query_db_lob_metrics(prev_db_file, lob, prev_q)
        
        # Calculate % changes
        pct_gross = 0.0
        if prev_metrics["gross_premium"] > 0:
            pct_gross = ((curr_metrics["gross_premium"] - prev_metrics["gross_premium"]) / prev_metrics["gross_premium"]) * 100.0
            
        pct_net = 0.0
        if prev_metrics["net_premium"] > 0:
            pct_net = ((curr_metrics["net_premium"] - prev_metrics["net_premium"]) / prev_metrics["net_premium"]) * 100.0
            
        pct_rein = 0.0
        if prev_metrics["rein_premium"] > 0:
            pct_rein = ((curr_metrics["rein_premium"] - prev_metrics["rein_premium"]) / prev_metrics["rein_premium"]) * 100.0
            
        pct_net_upr = 0.0
        if prev_metrics["net_upr"] > 0:
            pct_net_upr = ((curr_metrics["net_upr"] - prev_metrics["net_upr"]) / prev_metrics["net_upr"]) * 100.0
            
        pct_rein_upr = 0.0
        if prev_metrics["rein_upr"] > 0:
            pct_rein_upr = ((curr_metrics["rein_upr"] - prev_metrics["rein_upr"]) / prev_metrics["rein_upr"]) * 100.0

        lt_results.append({
            "lob": lob,
            # Old backward-compatibility keys
            "prev_premium": float(prev_metrics["gross_premium"]),
            "curr_premium": float(curr_metrics["gross_premium"]),
            "pct_change": float(pct_gross),
            "prev_upr": float(prev_metrics["net_upr"]),
            "curr_upr": float(curr_metrics["net_upr"]),
            # New detailed keys
            "prev_gross_premium": float(prev_metrics["gross_premium"]),
            "curr_gross_premium": float(curr_metrics["gross_premium"]),
            "pct_gross_premium": float(pct_gross),
            "prev_net_premium": float(prev_metrics["net_premium"]),
            "curr_net_premium": float(curr_metrics["net_premium"]),
            "pct_net_premium": float(pct_net),
            "prev_rein_premium": float(prev_metrics["rein_premium"]),
            "curr_rein_premium": float(curr_metrics["rein_premium"]),
            "pct_rein_premium": float(pct_rein),
            "prev_net_upr": float(prev_metrics["net_upr"]),
            "curr_net_upr": float(curr_metrics["net_upr"]),
            "pct_net_upr": float(pct_net_upr),
            "prev_rein_upr": float(prev_metrics["rein_upr"]),
            "curr_rein_upr": float(curr_metrics["rein_upr"]),
            "pct_rein_upr": float(pct_rein_upr)
        })

    return {
        "ok": True,
        "current_quarter": curr_q,
        "previous_quarter": prev_q,
        "short_term": st_results,
        "long_term": lt_results
    }

@router.get("/detail")
def get_premium_analysis_detail(day: int, month: int, year: int, lob: str, db: Session = Depends(get_db)):
    """Calculate the Premium & UPR composition (New vs Existing) for the selected LOB."""
    analysis_date = datetime.date(year, month, day)
    curr_q, prev_q = get_quarters_for_date(analysis_date)
    curr_qid = get_quarter_id_from_str(curr_q)
    prev_qid = get_quarter_id_from_str(prev_q)

    ty_gia = load_ty_gia()

    # Load current quarter valid rows
    curr_df = compute_lob_rows(lob, curr_qid, analysis_date, ty_gia)
    curr_valid = curr_df[curr_df["Quarter"] == curr_q].copy()

    # Load previous quarter valid rows
    prev_param = db.query(models.AppParameter).filter(models.AppParameter.quarter_id == prev_qid).first()
    if prev_param:
        prev_date = datetime.date(prev_param.year, prev_param.month, prev_param.day)
    else:
        prev_q_part, prev_y_part = prev_q.split("/")
        if prev_q_part == "Q1":
            prev_date = datetime.date(int(prev_y_part), 3, 31)
        elif prev_q_part == "Q2":
            prev_date = datetime.date(int(prev_y_part), 6, 30)
        elif prev_q_part == "Q3":
            prev_date = datetime.date(int(prev_y_part), 9, 30)
        else:
            prev_date = datetime.date(int(prev_y_part), 12, 31)

    prev_df = compute_lob_rows(lob, prev_qid, prev_date, ty_gia)
    prev_valid = prev_df[prev_df["Quarter"] == prev_q]

    # Create the set of valid policy numbers from the previous quarter
    prev_policy_set = set(prev_valid["Policy_Number"].dropna().unique())

    # Map New vs Existing on current quarter
    if not curr_valid.empty:
        curr_valid["Type"] = np.where(curr_valid["Policy_Number"].isin(prev_policy_set), "Existing", "New")
    else:
        curr_valid["Type"] = pd.Series(dtype=str)

    # Initialize totals
    premium_new_amount = 0.0
    premium_old_amount = 0.0
    upr_new_amount = 0.0
    upr_old_amount = 0.0

    premium_new_count = 0
    premium_old_count = 0
    upr_new_count = 0
    upr_old_count = 0

    if not curr_valid.empty:
        policy_sums = curr_valid.groupby(["Policy_Number", "Type"], as_index=False).agg({
            "Premium": "sum",
            "UPR": "sum"
        })

        premium_new_amount = float(curr_valid[curr_valid["Type"] == "New"]["Premium"].sum())
        premium_old_amount = float(curr_valid[curr_valid["Type"] == "Existing"]["Premium"].sum())
        upr_new_amount = float(curr_valid[curr_valid["Type"] == "New"]["UPR"].sum())
        upr_old_amount = float(curr_valid[curr_valid["Type"] == "Existing"]["UPR"].sum())

        premium_new_count = int(policy_sums[(policy_sums["Type"] == "New") & (policy_sums["Premium"] > 0)]["Policy_Number"].nunique())
        premium_old_count = int(policy_sums[(policy_sums["Type"] == "Existing") & (policy_sums["Premium"] > 0)]["Policy_Number"].nunique())
        upr_new_count = int(policy_sums[(policy_sums["Type"] == "New") & (policy_sums["UPR"] > 0)]["Policy_Number"].nunique())
        upr_old_count = int(policy_sums[(policy_sums["Type"] == "Existing") & (policy_sums["UPR"] > 0)]["Policy_Number"].nunique())

    premium_total = premium_new_amount + premium_old_amount
    upr_total = upr_new_amount + upr_old_amount

    premium_new_pct = (premium_new_amount / premium_total * 100.0) if premium_total > 0 else 0.0
    premium_old_pct = (premium_old_amount / premium_total * 100.0) if premium_total > 0 else 0.0

    upr_new_pct = (upr_new_amount / upr_total * 100.0) if upr_total > 0 else 0.0
    upr_old_pct = (upr_old_amount / upr_total * 100.0) if upr_total > 0 else 0.0

    stats_data = [
        {"class": "Premium (Amount)", "val": premium_total},
        {"class": "UPR (Amount)", "val": upr_total},
        {"class": "Premium - Existing Policies (Amount)", "val": premium_old_amount},
        {"class": "Premium - New Policies (Amount)", "val": premium_new_amount},
        {"class": "UPR - Existing Policies (Amount)", "val": upr_old_amount},
        {"class": "UPR - New Policies (Amount)", "val": upr_new_amount},
        {"class": "Premium - Existing Policies (Count)", "val": premium_old_count},
        {"class": "Premium - New Policies (Count)", "val": premium_new_count},
        {"class": "UPR - Existing Policies (Count)", "val": upr_old_count},
        {"class": "UPR - New Policies (Count)", "val": upr_new_count}
    ]

    return {
        "ok": True,
        "lob": lob,
        "pie_premium": {
            "new_pct": float(premium_new_pct),
            "old_pct": float(premium_old_pct),
            "new_val": float(premium_new_amount),
            "old_val": float(premium_old_amount)
        },
        "pie_upr": {
            "new_pct": float(upr_new_pct),
            "old_pct": float(upr_old_pct),
            "new_val": float(upr_new_amount),
            "old_val": float(upr_old_amount)
        },
        "stats": stats_data
    }

@router.get("/download-excel")
def download_premium_analysis_excel(day: int, month: int, year: int, db: Session = Depends(get_db)):
    """Export Premium & UPR overview tables side-by-side as a styled Excel sheet."""
    overview_data = get_premium_analysis_overview(day, month, year, db)
    
    curr_q = overview_data["current_quarter"]
    prev_q = overview_data["previous_quarter"]
    st_list = overview_data["short_term"]
    lt_list = overview_data["long_term"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Overview Analysis"

    font_title = Font(name="Calibri", size=16, bold=True, color="1E3A8A")
    font_subtitle = Font(name="Calibri", size=11, italic=True, color="4B5563")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_total = Font(name="Calibri", size=11, bold=True, color="111827")
    font_data = Font(name="Calibri", size=11, color="111827")
    
    fill_header = PatternFill("solid", fgColor="1E3A8A")
    fill_total = PatternFill("solid", fgColor="F3F4F6")
    fill_green = PatternFill("solid", fgColor="D1FAE5")
    fill_red = PatternFill("solid", fgColor="FEE2E2")

    border_thin = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    border_total = Border(
        top=Side(style='thin', color='9CA3AF'),
        bottom=Side(style='double', color='111827')
    )

    ws.merge_cells("A1:O1")
    ws["A1"] = "Premium & UPR Overview Analysis Report"
    ws["A1"].font = font_title
    
    ws.merge_cells("A2:O2")
    ws["A2"] = f"Analysis Date: {day:02d}/{month:02d}/{year}  |  Current Quarter: {curr_q} vs Previous Quarter: {prev_q}"
    ws["A2"].font = font_subtitle
    
    ws.merge_cells("A4:G4")
    ws["A4"] = "SHORT-TERM LINES OF BUSINESS"
    ws["A4"].font = Font(name="Calibri", size=12, bold=True, color="1E3A8A")
    
    ws.merge_cells("I4:O4")
    ws["I4"] = "LONG-TERM LINES OF BUSINESS"
    ws["I4"].font = Font(name="Calibri", size=12, bold=True, color="1E3A8A")

    headers_st = [
        "LOB", f"Premium {prev_q}", f"Premium {curr_q}", "% Change Premium",
        f"UPR {prev_q}", f"UPR {curr_q}", "% Change UPR"
    ]
    headers_lt = [
        "LOB", f"Premium {prev_q}", f"Premium {curr_q}", "% Change Premium",
        f"UPR {prev_q}", f"UPR {curr_q}", "% Change UPR"
    ]

    for col_idx, header in enumerate(headers_st, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers_lt, start=9):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    max_len = max(len(st_list), len(lt_list))
    for i in range(max_len):
        row_idx = 6 + i
        
        if i < len(st_list):
            item = st_list[i]
            ws.cell(row=row_idx, column=1, value=item["lob"]).font = font_data
            
            c_prev_p = ws.cell(row=row_idx, column=2, value=item["prev_premium"])
            c_prev_p.number_format = "#,##0"
            c_prev_p.font = font_data
            
            c_curr_p = ws.cell(row=row_idx, column=3, value=item["curr_premium"])
            c_curr_p.number_format = "#,##0"
            c_curr_p.font = font_data
            
            c_pct = ws.cell(row=row_idx, column=4, value=float(item["pct_change"]) / 100.0)
            c_pct.number_format = "0.00%"
            c_pct.font = font_data
            if item["pct_change"] > 0:
                c_pct.fill = fill_green
            elif item["pct_change"] < 0:
                c_pct.fill = fill_red
                
            c_prev_u = ws.cell(row=row_idx, column=5, value=item["prev_upr"])
            c_prev_u.number_format = "#,##0"
            c_prev_u.font = font_data
            
            c_curr_u = ws.cell(row=row_idx, column=6, value=item["curr_upr"])
            c_curr_u.number_format = "#,##0"
            c_curr_u.font = font_data
            
            c_pct_upr = ws.cell(row=row_idx, column=7, value=float(item["pct_upr"]) / 100.0)
            c_pct_upr.number_format = "0.00%"
            c_pct_upr.font = font_data
            if item["pct_upr"] > 0:
                c_pct_upr.fill = fill_green
            elif item["pct_upr"] < 0:
                c_pct_upr.fill = fill_red
            
            for c in range(1, 8):
                ws.cell(row=row_idx, column=c).border = border_thin

        if i < len(lt_list):
            item = lt_list[i]
            ws.cell(row=row_idx, column=9, value=item["lob"]).font = font_data
            
            c_prev_p = ws.cell(row=row_idx, column=10, value=item["prev_premium"])
            c_prev_p.number_format = "#,##0"
            c_prev_p.font = font_data
            
            c_curr_p = ws.cell(row=row_idx, column=11, value=item["curr_premium"])
            c_curr_p.number_format = "#,##0"
            c_curr_p.font = font_data
            
            c_pct = ws.cell(row=row_idx, column=12, value=float(item["pct_change"]) / 100.0)
            c_pct.number_format = "0.00%"
            c_pct.font = font_data
            if item["pct_change"] > 0:
                c_pct.fill = fill_green
            elif item["pct_change"] < 0:
                c_pct.fill = fill_red
                
            c_prev_u = ws.cell(row=row_idx, column=13, value=item["prev_upr"])
            c_prev_u.number_format = "#,##0"
            c_prev_u.font = font_data
            
            c_curr_u = ws.cell(row=row_idx, column=14, value=item["curr_upr"])
            c_curr_u.number_format = "#,##0"
            c_curr_u.font = font_data
            
            c_pct_upr = ws.cell(row=row_idx, column=15, value=float(item["pct_upr"]) / 100.0)
            c_pct_upr.number_format = "0.00%"
            c_pct_upr.font = font_data
            if item["pct_upr"] > 0:
                c_pct_upr.fill = fill_green
            elif item["pct_upr"] < 0:
                c_pct_upr.fill = fill_red
            
            for c in range(9, 16):
                ws.cell(row=row_idx, column=c).border = border_thin

    total_row_idx = 6 + max_len
    
    # ST Totals
    ws.cell(row=total_row_idx, column=1, value="Total").font = font_total
    ws.cell(row=total_row_idx, column=1).fill = fill_total
    
    sum_prev_p = f"=SUM(B6:B{total_row_idx-1})"
    c_tot_prev_p = ws.cell(row=total_row_idx, column=2, value=sum_prev_p)
    c_tot_prev_p.number_format = "#,##0"
    c_tot_prev_p.font = font_total
    c_tot_prev_p.fill = fill_total
    
    sum_curr_p = f"=SUM(C6:C{total_row_idx-1})"
    c_tot_curr_p = ws.cell(row=total_row_idx, column=3, value=sum_curr_p)
    c_tot_curr_p.number_format = "#,##0"
    c_tot_curr_p.font = font_total
    c_tot_curr_p.fill = fill_total
    
    formula_pct = f"=IF(B{total_row_idx}>0,(C{total_row_idx}-B{total_row_idx})/B{total_row_idx},0)"
    c_tot_pct = ws.cell(row=total_row_idx, column=4, value=formula_pct)
    c_tot_pct.number_format = "0.00%"
    c_tot_pct.font = font_total
    c_tot_pct.fill = fill_total
    
    sum_prev_u = f"=SUM(E6:E{total_row_idx-1})"
    c_tot_prev_u = ws.cell(row=total_row_idx, column=5, value=sum_prev_u)
    c_tot_prev_u.number_format = "#,##0"
    c_tot_prev_u.font = font_total
    c_tot_prev_u.fill = fill_total
    
    sum_curr_u = f"=SUM(F6:F{total_row_idx-1})"
    c_tot_curr_u = ws.cell(row=total_row_idx, column=6, value=sum_curr_u)
    c_tot_curr_u.number_format = "#,##0"
    c_tot_curr_u.font = font_total
    c_tot_curr_u.fill = fill_total
    
    formula_pct_upr = f"=IF(E{total_row_idx}>0,(F{total_row_idx}-E{total_row_idx})/E{total_row_idx},0)"
    c_tot_pct_upr = ws.cell(row=total_row_idx, column=7, value=formula_pct_upr)
    c_tot_pct_upr.number_format = "0.00%"
    c_tot_pct_upr.font = font_total
    c_tot_pct_upr.fill = fill_total
    
    for c in range(1, 8):
        ws.cell(row=total_row_idx, column=c).border = border_total

    # LT Totals
    ws.cell(row=total_row_idx, column=9, value="Total").font = font_total
    ws.cell(row=total_row_idx, column=9).fill = fill_total
    
    sum_prev_p_lt = f"=SUM(J6:J{total_row_idx-1})"
    c_tot_prev_p_lt = ws.cell(row=total_row_idx, column=10, value=sum_prev_p_lt)
    c_tot_prev_p_lt.number_format = "#,##0"
    c_tot_prev_p_lt.font = font_total
    c_tot_prev_p_lt.fill = fill_total
    
    sum_curr_p_lt = f"=SUM(K6:K{total_row_idx-1})"
    c_tot_curr_p_lt = ws.cell(row=total_row_idx, column=11, value=sum_curr_p_lt)
    c_tot_curr_p_lt.number_format = "#,##0"
    c_tot_curr_p_lt.font = font_total
    c_tot_curr_p_lt.fill = fill_total
    
    formula_pct_lt = f"=IF(J{total_row_idx}>0,(K{total_row_idx}-J{total_row_idx})/J{total_row_idx},0)"
    c_tot_pct_lt = ws.cell(row=total_row_idx, column=12, value=formula_pct_lt)
    c_tot_pct_lt.number_format = "0.00%"
    c_tot_pct_lt.font = font_total
    c_tot_pct_lt.fill = fill_total
    
    sum_prev_u_lt = f"=SUM(M6:M{total_row_idx-1})"
    c_tot_prev_u_lt = ws.cell(row=total_row_idx, column=13, value=sum_prev_u_lt)
    c_tot_prev_u_lt.number_format = "#,##0"
    c_tot_prev_u_lt.font = font_total
    c_tot_prev_u_lt.fill = fill_total
    
    sum_curr_u_lt = f"=SUM(N6:N{total_row_idx-1})"
    c_tot_curr_u_lt = ws.cell(row=total_row_idx, column=14, value=sum_curr_u_lt)
    c_tot_curr_u_lt.number_format = "#,##0"
    c_tot_curr_u_lt.font = font_total
    c_tot_curr_u_lt.fill = fill_total
    
    formula_pct_upr_lt = f"=IF(M{total_row_idx}>0,(N{total_row_idx}-M{total_row_idx})/M{total_row_idx},0)"
    c_tot_pct_upr_lt = ws.cell(row=total_row_idx, column=15, value=formula_pct_upr_lt)
    c_tot_pct_upr_lt.number_format = "0.00%"
    c_tot_pct_upr_lt.font = font_total
    c_tot_pct_upr_lt.fill = fill_total
    
    for c in range(9, 16):
        ws.cell(row=total_row_idx, column=c).border = border_total

    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.row in [1, 2, 4]:
                continue
            val_str = str(cell.value or "")
            if cell.number_format == "#,##0" and isinstance(cell.value, (int, float)):
                val_str = f"{int(cell.value):,}"
            elif cell.number_format == "0.00%" and isinstance(cell.value, (int, float)):
                val_str = f"{cell.value * 100:.2f}%"
            max_len = max(max_len, len(val_str))
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    ws.column_dimensions["H"].width = 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    filename = f"Premium_Analysis_Overview_{year}{month:02d}{day:02d}.xlsx"
    
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
