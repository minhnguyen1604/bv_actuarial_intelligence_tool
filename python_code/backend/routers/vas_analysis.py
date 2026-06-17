import io
import os
import re
import shutil
import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, BackgroundTasks
from fastapi.responses import StreamingResponse, JSONResponse, FileResponse
from pydantic import BaseModel
from typing import List, Optional
import uuid

from services import vas_calculator

router = APIRouter(
    prefix="/api/vas",
    tags=["vas_analysis"],
)

UPLOAD_DIR = os.path.join("uploads", "vas")
os.makedirs(UPLOAD_DIR, exist_ok=True)

class CheckRequest(BaseModel):
    file_path: str
    sheet_result: str
    sheet_dtbt: str

class CalculateRequest(BaseModel):
    file_path: str
    sheet_result: str
    sheet_dtbt: str
    quarter_id: str  # e.g., "2025Q1"

class UpdateRequest(BaseModel):
    quarter_id: str
    confirm_replace: bool = False

# numpy NaN/inf encoder helper
def sanitize_numpy(obj):
    import math
    if obj is None:
        return None
    type_name = type(obj).__name__
    if type_name in ("int64", "int32", "int16", "int8", "uint64", "uint32", "uint16", "uint8"):
        return int(obj)
    if type_name in ("float64", "float32", "float16"):
        if math.isnan(float(obj)) or math.isinf(float(obj)):
            return None
        return float(obj)
    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj
    if isinstance(obj, dict):
        return {str(k): sanitize_numpy(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [sanitize_numpy(i) for i in obj]
    return obj

@router.post("/upload")
async def upload_file(file: UploadFile = File(...), quarter_id: str = Form(...)):
    if not file.filename.lower().endswith(('.xls', '.xlsx', '.xlsb', '.xlsm')):
        raise HTTPException(status_code=400, detail="Only Excel files are allowed.")
        
    quarter_dir = os.path.join(UPLOAD_DIR, quarter_id)
    os.makedirs(quarter_dir, exist_ok=True)
    file_path = os.path.join(quarter_dir, file.filename)
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    # Read sheet names using openpyxl
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read sheets: {str(e)}")
        
    # Get suggested sheets using fuzzy logic
    best_res, best_dt = vas_calculator.find_best_sheets(file_path)
    
    return {
        "file_path": file_path,
        "file_name": file.filename,
        "sheets": sheets,
        "recommended_result_sheet": best_res,
        "recommended_dtbt_sheet": best_dt
    }

@router.post("/check")
def check_sheets(req: CheckRequest):
    if not os.path.exists(req.file_path):
        raise HTTPException(status_code=404, detail="Uploaded file not found.")
        
    try:
        # Check sheet1 (Result) structure
        df1 = pd.read_excel(req.file_path, sheet_name=req.sheet_result, dtype=str)
        sheet1_ok = vas_calculator.check_form(df1)
        
        # Check sheet2 (Doanh thu bồi thường) structure
        df2 = pd.read_excel(req.file_path, sheet_name=req.sheet_dtbt, header=None)
        try:
            vas_calculator.parse_dtbt_sheet(df2)
            sheet2_ok = True
        except Exception:
            sheet2_ok = False
        
        ok = sheet1_ok and sheet2_ok
        detail = ""
        if not ok:
            errs = []
            if not sheet1_ok:
                errs.append("Sheet 1 (Kết quả) không đúng định dạng dự phòng (UPR/OSC/IBNR/CAT).")
            if not sheet2_ok:
                errs.append("Sheet 2 (Doanh thu bồi thường) thiếu từ khóa tiêu đề hoặc dòng phân tách 'BT'.")
            detail = " ".join(errs)
            
        return {
            "ok": ok,
            "sheet1_ok": sheet1_ok,
            "sheet2_ok": sheet2_ok,
            "detail": detail
        }
    except Exception as e:
        return {
            "ok": False,
            "sheet1_ok": False,
            "sheet2_ok": False,
            "detail": f"Lỗi đọc file: {str(e)}"
        }

BACKGROUND_TASKS = {}

def async_calculate_vas(task_id: str, file_path: str, sheet_result: str, sheet_dtbt: str, quarter_id: str):
    try:
        BACKGROUND_TASKS[task_id]["progress"] = 10
        BACKGROUND_TASKS[task_id]["message"] = "Đang đọc các sheet từ file Excel..."
        
        # 1. Read Excel sheets
        df1 = pd.read_excel(file_path, sheet_name=sheet_result)
        df2 = pd.read_excel(file_path, sheet_name=sheet_dtbt, header=None)
        
        BACKGROUND_TASKS[task_id]["progress"] = 30
        BACKGROUND_TASKS[task_id]["message"] = "Đang phân tích cấu trúc dự phòng & doanh thu..."
        
        # 2. Parse sheet1 and sheet2
        df_joined_res = vas_calculator.parse_result_sheet(df1)
        dt_final, bt_final = vas_calculator.parse_dtbt_sheet(df2)
        
        BACKGROUND_TASKS[task_id]["progress"] = 55
        BACKGROUND_TASKS[task_id]["message"] = "Đang ghép và so khớp dữ liệu quý trước..."
        
        # Rename to lowercase to match database columns
        dt_final = dt_final.rename(columns={"Written": "written"})
        bt_final = bt_final.rename(columns={"Paid": "paid"})
        
        # Join Written, Paid with reserves
        df_joined = pd.merge(dt_final, bt_final, on=["Line", "Type"], how="outer")
        df_joined = pd.merge(df_joined, df_joined_res, on=["Line", "Type"], how="outer")
        df_joined = df_joined.fillna(0.0)
        
        is_q1 = quarter_id.endswith("Q1")
        prev_q_n = vas_calculator.prev_quarter(quarter_id)
        
        # 3. Quarterly changes (Theo quý)
        prev_quarter_df = vas_calculator.get_history_by_quarter(prev_q_n, "Theo quý")
        df_compare_q = vas_calculator.calculate_quarterly_changes(df_joined, prev_quarter_df, is_q1)
        
        BACKGROUND_TASKS[task_id]["progress"] = 75
        BACKGROUND_TASKS[task_id]["message"] = "Đang tổng hợp số liệu lũy kế & 4Q gần nhất..."
        
        # 4. Cumulative changes (Lũy kế)
        prev_quarter_cum_df = vas_calculator.get_history_by_quarter(prev_q_n, "Lũy kế")
        df_now = vas_calculator.calculate_cumulative_changes(df_compare_q, prev_quarter_cum_df, is_q1)
        
        # 5. Trailing 12 months (4Q)
        last4_q = vas_calculator.get_last_4_quarters(quarter_id)
        # Fetch quarterly history for last 4 quarters
        history_list = []
        for q in last4_q:
            if q == quarter_id:
                history_list.append(df_compare_q.copy().rename(columns={"Line": "line", "Type": "type"}))
            else:
                hist_q = vas_calculator.get_history_by_quarter(q, "Theo quý")
                if not hist_q.empty:
                    history_list.append(hist_q)
        if len(history_list) > 0:
            history_q_df = pd.concat(history_list, ignore_index=True)
        else:
            history_q_df = pd.DataFrame()
            
        prev_q_4 = vas_calculator.prev_quarter(vas_calculator.prev_quarter(vas_calculator.prev_quarter(prev_q_n)))
        lag4_reserves_df = vas_calculator.get_history_by_quarter(prev_q_4, "Theo quý")
        
        df_4q = vas_calculator.calculate_trailing_12m(df_compare_q, history_q_df, lag4_reserves_df)
        
        BACKGROUND_TASKS[task_id]["progress"] = 90
        BACKGROUND_TASKS[task_id]["message"] = "Đang hoàn tất tính toán tỷ lệ tài chính..."
        
        # 6. Merge all
        he = pd.concat([df_compare_q, df_now, df_4q], ignore_index=True)
        he = vas_calculator.compute_financial_ratios(he)
        
        # Set keys
        he["nam"] = quarter_id
        he["nam_lk"] = quarter_id[:4]
        he["quy"] = quarter_id[4:]
        
        # Save temporary result to uploads/vas/<quarter_id>/temp_result.parquet
        temp_dir = os.path.dirname(file_path)
        temp_path = os.path.join(temp_dir, "temp_result.parquet")
        he.to_parquet(temp_path, index=False)
        
        BACKGROUND_TASKS[task_id]["progress"] = 100
        BACKGROUND_TASKS[task_id]["message"] = "Đã hoàn thành ghép và tính toán!"
        BACKGROUND_TASKS[task_id]["status"] = "success"
        BACKGROUND_TASKS[task_id]["result"] = he.to_dict(orient="records")
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        BACKGROUND_TASKS[task_id]["status"] = "failed"
        BACKGROUND_TASKS[task_id]["error"] = str(e)
        BACKGROUND_TASKS[task_id]["message"] = f"Lỗi tính toán: {str(e)}"

@router.post("/calculate")
def calculate_vas(req: CalculateRequest, bg_tasks: BackgroundTasks):
    if not os.path.exists(req.file_path):
        raise HTTPException(status_code=404, detail="Uploaded file not found.")
        
    task_id = str(uuid.uuid4())
    BACKGROUND_TASKS[task_id] = {
        "status": "running",
        "progress": 0,
        "message": "Đang khởi chạy tác vụ tính toán...",
        "result": None,
        "error": None
    }
    
    bg_tasks.add_task(async_calculate_vas, task_id, req.file_path, req.sheet_result, req.sheet_dtbt, req.quarter_id)
    return {"task_id": task_id, "status": "running"}

@router.get("/task-status/{task_id}")
def get_task_status(task_id: str):
    if task_id not in BACKGROUND_TASKS:
        raise HTTPException(status_code=404, detail="Task ID not found.")
    
    task = BACKGROUND_TASKS[task_id]
    response_data = {
        "status": task["status"],
        "progress": task["progress"],
        "message": task["message"],
        "error": task["error"]
    }
    if task["status"] == "success":
        response_data["result"] = sanitize_numpy(task["result"])
        
    return JSONResponse(content=response_data)

@router.post("/update")
def update_history(req: UpdateRequest):
    temp_path = os.path.join(UPLOAD_DIR, req.quarter_id, "temp_result.parquet")
    if not os.path.exists(temp_path):
        raise HTTPException(status_code=400, detail="No calculated temp data found. Run calculate first.")
        
    # Check if duplicate in SQLite
    if not os.path.exists(vas_calculator.DB_PATH):
        pass
    else:
        conn = sqlite3.connect(vas_calculator.DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM vas_history WHERE nam = ?", (req.quarter_id,))
        exists = cursor.fetchone()[0] > 0
        conn.close()
        
        if exists and not req.confirm_replace:
            return {
                "status": "duplicate",
                "message": f"Dữ liệu cho {req.quarter_id} đã tồn tại trong hệ thống. Bạn có muốn thay thế không?"
            }
            
    try:
        df = pd.read_parquet(temp_path)
        vas_calculator.save_to_history(df, req.quarter_id)
        
        # Copy original uploaded file to physical archive folder
        quarter_dir = os.path.join(UPLOAD_DIR, req.quarter_id)
        if os.path.exists(quarter_dir):
            excel_files = [f for f in os.listdir(quarter_dir) if f.lower().endswith(('.xls', '.xlsx', '.xlsb', '.xlsm'))]
            if excel_files:
                src_path = os.path.join(quarter_dir, excel_files[0])
                ext = os.path.splitext(excel_files[0])[1]
                archive_dir = os.path.join("data", "archive")
                os.makedirs(archive_dir, exist_ok=True)
                dest_path = os.path.join(archive_dir, f"VAS_Archive_{req.quarter_id}{ext}")
                shutil.copy2(src_path, dest_path)
                print(f"Archived original Excel file to: {dest_path}")
                
        return {
            "status": "success",
            "message": "Đã ghép dữ liệu và lưu trữ file gốc thành công!"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi lưu trữ: {str(e)}")

@router.get("/history")
def get_history():
    if not os.path.exists(vas_calculator.DB_PATH):
        return []
        
    conn = sqlite3.connect(vas_calculator.DB_PATH)
    try:
        df = pd.read_sql_query("SELECT * FROM vas_history", conn)
        return JSONResponse(content=sanitize_numpy(df.to_dict(orient="records")))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@router.get("/compare")
def get_comparison(quy_luy_ke: str, type_sel: str, metric: str, start_year: int):
    if not os.path.exists(vas_calculator.DB_PATH):
        raise HTTPException(status_code=404, detail="Database not initialized yet.")
        
    conn = sqlite3.connect(vas_calculator.DB_PATH)
    try:
        # Check standard columns mapping
        db_metric = metric.lower().replace("@", "reserve_").replace("/", "_").replace("(", "_").replace(")", "_").replace(" ", "_")
        if metric == "Sub total(earned)":
            db_metric = "sub_total_earned"
        elif metric == "Sub total(incurred)":
            db_metric = "sub_total_incurred"
        elif metric == "Paid/Written":
            db_metric = "paid_written"
        elif metric == "Incurred/Earned":
            db_metric = "incurred_earned"
        elif metric == "@OsC":
            db_metric = "reserve_osc"
        elif metric == "@UPR":
            db_metric = "reserve_upr"
        elif metric == "@IBNR":
            db_metric = "reserve_ibnr"
        elif metric == "@CAT":
            db_metric = "reserve_cat"
            
        # Section 1: compare metric over years for LOBs
        query1 = """
            SELECT line as Line, nam, {metric_col} as val 
            FROM vas_history 
            WHERE quy_luy_ke = ? AND type = ? AND CAST(nam_lk AS INTEGER) >= ?
        """.format(metric_col=db_metric)
        
        df1 = pd.read_sql_query(query1, conn, params=(quy_luy_ke, type_sel, start_year))
        if df1.empty:
            return {
                "table_data": [],
                "anomalies": {},
                "last_col": "",
                "type_comparison_table": [],
                "type_rules": {"recovery_issue": [], "retro_issue": []}
            }
            
        # Pivot wider: index=Line, columns=nam, values=val
        df_pivot = df1.pivot(index="Line", columns="nam", values="val").reset_index()
        # Sort Line using custom ordering
        line_order = [
            "Total", "Healthcare", "Travel", "Healthcare + Travel", "Personal Accident",
            "HC & PA & Travel", "Motor Vehicles", "Engineering", "Fire and Misc.",
            "General Liability", "Cargo in transit", "Hull & PI", "Aviation & Oil", "Agriculture"
        ]
        df_pivot["Line"] = pd.Categorical(df_pivot["Line"], categories=line_order, ordered=True)
        df_pivot = df_pivot.sort_values("Line").reset_index(drop=True)
        # Ensure Line name is string
        df_pivot["Line"] = df_pivot["Line"].astype(str)
        
        # Scaling if currency (everything except ratios is multiplied by 1e6 in display)
        is_ratio = metric in ["Paid/Written", "Incurred/Earned"]
        if not is_ratio:
            for col in df_pivot.columns:
                if col != "Line":
                    df_pivot[col] = df_pivot[col] * 1e6
                    
        # Anomaly detection
        anomalies_res = vas_calculator.detect_anomalies(df_pivot, threshold_yoy=0.3, threshold_z=3.0)
        
        # Section 2: compare metric across types for latest quarter
        # Find latest quarter in db
        cursor = conn.cursor()
        cursor.execute("SELECT MAX(nam) FROM vas_history")
        latest_period = cursor.fetchone()[0]
        
        query2 = """
            SELECT line as Line, type, {metric_col} as val
            FROM vas_history
            WHERE quy_luy_ke = ? AND nam = ?
        """.format(metric_col=db_metric)
        df2 = pd.read_sql_query(query2, conn, params=(quy_luy_ke, latest_period))
        
        df_type_pivot = pd.DataFrame()
        type_rules = {"recovery_issue": [], "retro_issue": []}
        
        if not df2.empty:
            df_type_pivot = df2.pivot(index="Line", columns="type", values="val").reset_index()
            df_type_pivot["Line"] = pd.Categorical(df_type_pivot["Line"], categories=line_order, ordered=True)
            df_type_pivot = df_type_pivot.sort_values("Line").reset_index(drop=True)
            df_type_pivot["Line"] = df_type_pivot["Line"].astype(str)
            
            # Apply color check limits
            type_rules = vas_calculator.check_structure_rules(df_type_pivot)
            
            # Scale currency values
            if not is_ratio:
                for col in df_type_pivot.columns:
                    if col != "Line":
                        df_type_pivot[col] = df_type_pivot[col] * 1e6
                        
        return JSONResponse(content=sanitize_numpy({
            "table_data": df_pivot.to_dict(orient="records"),
            "anomalies": anomalies_res["anomalies"],
            "last_col": anomalies_res["last_col"],
            "type_comparison_table": df_type_pivot.to_dict(orient="records") if not df_type_pivot.empty else [],
            "type_rules": type_rules,
            "latest_period": latest_period
        }))
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@router.get("/detail")
def get_lob_detail(line: str, year_quarter: str, quy_luy_ke: str):
    if not os.path.exists(vas_calculator.DB_PATH):
        raise HTTPException(status_code=404, detail="Database not initialized yet.")
        
    conn = sqlite3.connect(vas_calculator.DB_PATH)
    try:
        # 1. Detailed comparison table of the LOB for the selected quarter across Types
        query1 = """
            SELECT type, 
                   osc*1e6 as OsC_trich_them, 
                   upr*1e6 as UPR_trich_them, 
                   reserve_osc*1e6 as reserve_osc, 
                   reserve_upr*1e6 as reserve_upr, 
                   written*1e6 as Written, 
                   paid*1e6 as Paid, 
                   sub_total_earned*1e6 as sub_total_earned, 
                   sub_total_incurred*1e6 as sub_total_incurred, 
                   incurred_earned as incurred_earned
            FROM vas_history 
            WHERE line = ? AND nam = ? AND quy_luy_ke = ?
        """
        df1 = pd.read_sql_query(query1, conn, params=(line, year_quarter, quy_luy_ke))
        
        # Pivot longer then wider
        detail_records = []
        consistency_checks = {
            "abs_violation": [],
            "sign_violation": [],
            "written_ratio": 0.0,
            "paid_ratio": 0.0
        }
        
        if not df1.empty:
            # Reorder types
            type_order = ["DIRECT", "RECOVERY", "INWARD", "RETROCESSION", "NET"]
            df1["type"] = pd.Categorical(df1["type"], categories=type_order, ordered=True)
            df1 = df1.sort_values("type").reset_index(drop=True)
            
            # Melt
            melted = df1.melt(id_vars=["type"], var_name="Chỉ tiêu", value_name="Giá trị")
            # Map metric names for display
            metric_map = {
                "OsC_trich_them": "OsC_trich_them",
                "UPR_trich_them": "UPR_trich_them",
                "reserve_osc": "@OsC",
                "reserve_upr": "@UPR",
                "Written": "Written",
                "Paid": "Paid",
                "sub_total_earned": "Sub total(earned)",
                "sub_total_incurred": "Sub total(incurred)",
                "incurred_earned": "Incurred/Earned"
            }
            melted["Chỉ tiêu"] = melted["Chỉ tiêu"].map(metric_map)
            
            # Pivot wider
            df_pivot = melted.pivot(index="Chỉ tiêu", columns="type", values="Giá trị").reset_index()
            # Custom ordering for metrics
            metric_order = [
                "OsC_trich_them", "UPR_trich_them", "@OsC", "@UPR",
                "Written", "Paid", "Sub total(earned)", "Sub total(incurred)", "Incurred/Earned"
            ]
            df_pivot["Chỉ tiêu"] = pd.Categorical(df_pivot["Chỉ tiêu"], categories=metric_order, ordered=True)
            df_pivot = df_pivot.sort_values("Chỉ tiêu").reset_index(drop=True)
            df_pivot["Chỉ tiêu"] = df_pivot["Chỉ tiêu"].astype(str)
            
            detail_records = df_pivot.to_dict(orient="records")
            
            # Run consistency checks on DIRECT vs RECOVERY
            consistency_checks = vas_calculator.check_direct_recovery_consistency(df_pivot)
            
        # 2. Historical loss ratios (Direct vs Net) for the last 9 quarters
        # direct
        query2 = """
            SELECT nam as quarter, sub_total_earned*1e6 as Earned, sub_total_incurred*1e6 as Incurred, incurred_earned as Loss_ratio
            FROM vas_history
            WHERE line = ? AND type = 'DIRECT' AND quy_luy_ke = ?
            ORDER BY nam ASC
        """
        df_direct = pd.read_sql_query(query2, conn, params=(line, quy_luy_ke))
        
        # net
        query3 = """
            SELECT nam as quarter, sub_total_earned*1e6 as Earned, sub_total_incurred*1e6 as Incurred, incurred_earned as Loss_ratio
            FROM vas_history
            WHERE line = ? AND type = 'NET' AND quy_luy_ke = ?
            ORDER BY nam ASC
        """
        df_net = pd.read_sql_query(query3, conn, params=(line, quy_luy_ke))
        
        def format_loss_ratio_history(df_hist):
            if df_hist.empty:
                return []
            df_hist = df_hist.tail(9)  # Keep last 9 quarters
            # Melt
            melted = df_hist.melt(id_vars=["quarter"], var_name="Chỉ tiêu", value_name="Giá trị")
            # Pivot wider
            df_pivot = melted.pivot(index="Chỉ tiêu", columns="quarter", values="Giá trị").reset_index()
            # Sort indicators
            df_pivot["Chỉ tiêu"] = pd.Categorical(df_pivot["Chỉ tiêu"], categories=["Earned", "Incurred", "Loss_ratio"], ordered=True)
            df_pivot = df_pivot.sort_values("Chỉ tiêu").reset_index(drop=True)
            df_pivot["Chỉ tiêu"] = df_pivot["Chỉ tiêu"].astype(str)
            return df_pivot.to_dict(orient="records")
            
        history_direct = format_loss_ratio_history(df_direct)
        history_net = format_loss_ratio_history(df_net)
        
        return JSONResponse(content=sanitize_numpy({
            "detail_table": detail_records,
            "consistency_checks": consistency_checks,
            "history_direct": history_direct,
            "history_net": history_net
        }))
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@router.get("/download-all-data")
def download_all_data():
    """Download entire vas_history database as an Excel file"""
    if not os.path.exists(vas_calculator.DB_PATH):
        raise HTTPException(status_code=404, detail="Database not initialized yet.")
        
    conn = sqlite3.connect(vas_calculator.DB_PATH)
    try:
        df = pd.read_sql_query("SELECT * FROM vas_history", conn)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "VAS_History"
        
        headers = list(df.columns)
        header_fill = PatternFill("solid", fgColor="0056A3")  # BaoViet Blue
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            
        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
            for col_idx, col_name in enumerate(headers, start=1):
                val = row[col_name]
                if pd.isna(val):
                    val = None
                ws.cell(row=row_idx, column=col_idx, value=val)
                
        # Auto size columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        from urllib.parse import quote
        filename_xlsx = f"VAS_History_Data_{date.today().isoformat()}.xlsx"
        encoded_name = quote(filename_xlsx, safe="")
        
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=\"{filename_xlsx}\"; filename*=UTF-8''{encoded_name}"},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

# New Endpoints for Data Rollback, Archived File Download, and Quarters List

class DeleteRequest(BaseModel):
    quarter_id: str
    password: str

@router.get("/quarters")
def get_quarters():
    if not os.path.exists(vas_calculator.DB_PATH):
        return []
    
    conn = sqlite3.connect(vas_calculator.DB_PATH)
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT nam FROM vas_history ORDER BY nam DESC")
        quarters = [row[0] for row in cursor.fetchall() if row[0]]
        return quarters
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@router.post("/delete")
def delete_quarter(req: DeleteRequest):
    if req.password != "68686868":
        raise HTTPException(status_code=403, detail="Mật khẩu bảo mật không chính xác!")
        
    if not os.path.exists(vas_calculator.DB_PATH):
        raise HTTPException(status_code=404, detail="Không tìm thấy cơ sở dữ liệu.")
        
    conn = sqlite3.connect(vas_calculator.DB_PATH)
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM vas_history WHERE nam = ?", (req.quarter_id,))
        exists = cursor.fetchone()[0] > 0
        if not exists:
            raise HTTPException(status_code=404, detail=f"Không tìm thấy dữ liệu cho kỳ {req.quarter_id}.")
            
        cursor.execute("DELETE FROM vas_history WHERE nam = ?", (req.quarter_id,))
        conn.commit()
        
        # Also delete physical archived Excel file if it exists
        archive_dir = os.path.join("data", "archive")
        if os.path.exists(archive_dir):
            for f in os.listdir(archive_dir):
                if f.startswith(f"VAS_Archive_{req.quarter_id}."):
                    file_to_del = os.path.join(archive_dir, f)
                    try:
                        os.remove(file_to_del)
                        print(f"Deleted archived file: {file_to_del}")
                    except Exception:
                        pass
                        
        return {"status": "success", "message": f"Đã xóa thành công toàn bộ số liệu và tệp nguồn kỳ {req.quarter_id}!"}
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@router.get("/download-archive/{quarter_id}")
def download_archive(quarter_id: str):
    archive_dir = os.path.join("data", "archive")
    if not os.path.exists(archive_dir):
        raise HTTPException(status_code=404, detail="Thư mục lưu trữ tệp gốc trống.")
        
    archived_file = None
    for f in os.listdir(archive_dir):
        if f.startswith(f"VAS_Archive_{quarter_id}."):
            archived_file = f
            break
            
    if not archived_file:
        raise HTTPException(status_code=404, detail=f"Không tìm thấy tệp Excel gốc được lưu trữ cho kỳ {quarter_id}.")
        
    file_path = os.path.join(archive_dir, archived_file)
    from urllib.parse import quote
    encoded_name = quote(archived_file, safe="")
    
    return FileResponse(
        path=file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=archived_file,
        headers={"Content-Disposition": f"attachment; filename=\"{archived_file}\"; filename*=UTF-8''{encoded_name}"}
    )
