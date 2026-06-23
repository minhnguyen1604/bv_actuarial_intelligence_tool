import os
import re
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
from sqlalchemy.orm import Session

import models

# Directory definitions
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_ROOT = os.path.join(BASE_DIR, "..", "data")
OUTPUT_EXCEL_ROOT = os.path.join(BASE_DIR, "..", "output_excel")

# Ensure folders exist
os.makedirs(OUTPUT_EXCEL_ROOT, exist_ok=True)

import tempfile
TEMP_DIR = os.path.join(DATA_ROOT, "tmp")
os.makedirs(TEMP_DIR, exist_ok=True)
tempfile.tempdir = TEMP_DIR

def int2col(col_idx: int) -> str:
    """Convert a 1-based column index to an Excel column letter."""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result

def load_ty_gia() -> pd.DataFrame:
    """Load the exchange rates dataframe directly from the SQLite database."""
    from database import SessionLocal
    db = SessionLocal()
    try:
        db_rates = db.query(models.FXRate).all()
    finally:
        db.close()

    cols = [
        'Thoi_gian', 'USD', 'EUR', 'AUD', 'CHF', 'CAD', 'CNY', 'DKK', 'GBP', 
        'HKD', 'INR', 'JPY', 'KRW', 'KWD', 'MYR', 'NOK', 'RUB', 'SAR', 'SEK', 
        'SGD', 'THB'
    ]

    if not db_rates:
        return pd.DataFrame(columns=cols)

    # Group by quarter_id
    import collections
    rates_by_quarter = collections.defaultdict(dict)
    for r in db_rates:
        rates_by_quarter[r.quarter_id][r.currency] = r.rate

    # Build rows
    rows = []
    for q_id, cur_map in rates_by_quarter.items():
        try:
            q_part, y_part = q_id.split("_")
            q_num = int(q_part[1]) # "Q1" -> 1
            year = int(y_part)
        except Exception:
            q_num = 0
            year = 0
            
        row = {"Thoi_gian": q_id.replace("_", "/"), "_sort_key": (year, q_num)}
        for col in cols:
            if col != "Thoi_gian":
                row[col] = cur_map.get(col, None)
        rows.append(row)

    # Sort rows chronologically
    rows.sort(key=lambda x: x["_sort_key"])

    # Remove temporary sort key
    for r in rows:
        del r["_sort_key"]

    return pd.DataFrame(rows, columns=cols)

def recalculate_excel_file(file_path: str):
    """No-op. Excel COM recalculation has been deprecated for performance."""
    pass

def calculate_st_summary_df(df: pd.DataFrame, ky_available: list[int], ty_gia: pd.DataFrame, dpnv_date: datetime.date, four_last_quarters: list[str]) -> pd.DataFrame:
    summary_rows = []
    
    ty_gia_dict_usd = dict(zip(ty_gia["Thoi_gian"].astype(str), ty_gia["USD"]))
    ty_gia_dict_eur = dict(zip(ty_gia["Thoi_gian"].astype(str), ty_gia["EUR"]))
    fallback_usd = ty_gia["USD"].iloc[-1] if not ty_gia.empty else 1.0
    fallback_eur = ty_gia["EUR"].iloc[-1] if not ty_gia.empty else 1.0
    
    dpnv = pd.to_datetime(dpnv_date)
    
    ret_col = "Ty_le_giu_lai_cua_BHBV" if "Ty_le_giu_lai_cua_BHBV" in df.columns else \
              ("Ty_le_giu_lai_cua_BHBV_checked" if "Ty_le_giu_lai_cua_BHBV_checked" in df.columns else None)
              
    for ky in ky_available:
        sh = f"Ky_phi{ky}"
        
        thang = pd.to_numeric(df[f"Ky_phi_{ky}_Ghi_Thang"], errors='coerce').fillna(0).astype(int)
        nam = pd.to_numeric(df[f"Ky_phi_{ky}_Ghi_Nam"], errors='coerce').fillna(0).astype(int)
        
        quy = np.where(thang == 0, 0, (thang - 1) // 3 + 1)
        quy_nam = "Q" + pd.Series(quy).astype(str) + "/" + pd.Series(nam).astype(str)
        quy_nam = np.where((quy == 0) | (nam == 0), "", quy_nam)
        
        rate_usd = pd.Series(quy_nam).map(ty_gia_dict_usd).fillna(fallback_usd).fillna(0.0).values
        rate_eur = pd.Series(quy_nam).map(ty_gia_dict_eur).fillna(fallback_eur).fillna(0.0).values
        
        vnd = pd.to_numeric(df[f"Ky_phi_{ky}_So_tien_VND"].astype(str).str.replace(",", "", regex=False), errors='coerce').fillna(0.0).values
        usd = pd.to_numeric(df[f"Ky_phi_{ky}_So_tien_USD"].astype(str).str.replace(",", "", regex=False), errors='coerce').fillna(0.0).values
        eur = pd.to_numeric(df[f"Ky_phi_{ky}_So_tien_EUR"].astype(str).str.replace(",", "", regex=False), errors='coerce').fillna(0.0).values
        
        phi_goc = vnd + usd * rate_usd + eur * rate_eur
        
        if ret_col:
            hi = pd.to_numeric(df[ret_col].astype(str).str.replace(",", "", regex=False), errors='coerce').fillna(0.0).values
            tile = np.where((hi == 0.0) | df[ret_col].isna(), 1.0, np.where(hi > 1.0, hi / 100.0, hi))
        else:
            tile = np.ones(len(df))
            
        phi_giulai = phi_goc * tile
        
        def parse_date_series(d_col, m_col, y_col):
            d_val = pd.to_numeric(df[d_col], errors='coerce')
            m_val = pd.to_numeric(df[m_col], errors='coerce')
            y_val = pd.to_numeric(df[y_col], errors='coerce')
            d_str = y_val.fillna(2000).astype(int).astype(str).str.zfill(4) + '-' + \
                    m_val.fillna(1).astype(int).astype(str).str.zfill(2) + '-' + \
                    d_val.fillna(1).astype(int).astype(str).str.zfill(2)
            parsed = pd.to_datetime(d_str, format='%Y-%m-%d', errors='coerce')
            parsed = parsed.where(d_val.notna() & m_val.notna() & y_val.notna(), pd.NaT)
            return parsed
            
        tu_date = parse_date_series("Thoi_han_bao_hiem_Tu_Ngay", "Thoi_han_bao_hiem_Tu_Thang", "Thoi_han_bao_hiem_Tu_Nam")
        den_date = parse_date_series("Thoi_han_bao_hiem_Den_Ngay", "Thoi_han_bao_hiem_Den_Thang", "Thoi_han_bao_hiem_Den_Nam")
        
        diff_days = (den_date - tu_date).dt.days
        dem_ngay = np.where(diff_days.isna(), 0, np.where(diff_days < 365, 1, 0))
        
        diff_dpnv_den = (dpnv - den_date).dt.days
        hethieuluc = np.where(diff_dpnv_den.isna(), 0, np.where(diff_dpnv_den >= 0, 1, 0))
        
        for q in four_last_quarters:
            mask_q = (quy_nam == q)
            
            phi_goc_sum = phi_goc[mask_q].sum()
            phi_giulai_sum = phi_giulai[mask_q].sum()
            
            mask_giam = mask_q & (dem_ngay == 1) & (hethieuluc == 1)
            giam_goc_sum = phi_goc[mask_giam].sum()
            giam_giulai_sum = phi_giulai[mask_giam].sum()
            giam_tai_sum = giam_goc_sum - giam_giulai_sum
            
            summary_rows.append({
                "Ky_phi": sh,
                "Quy": q,
                "Phi_bao_hiem_goc": phi_goc_sum,
                "Phi_bao_hiem_giu_lai": phi_giulai_sum,
                "Giam_phi_bao_hiem_goc": giam_goc_sum,
                "Giam_phi_bao_hiem_giu_lai": giam_giulai_sum,
                "Giam_phi_bao_hiem_tai": giam_tai_sum
            })
            
    return pd.DataFrame(summary_rows)

def calculate_vietjet_summary_df(df: pd.DataFrame, four_last_quarters: list[str]) -> pd.DataFrame:
    summary_rows = []
    
    thang = pd.to_numeric(df["Thang_phat_sinh_doanh_thu"], errors='coerce').fillna(0).astype(int)
    nam = pd.to_numeric(df["Nam_phat_sinh_doanh_thu"], errors='coerce').fillna(0).astype(int)
    
    quy = np.where(thang == 0, 0, np.where(thang == 3, 2, (thang - 1) // 3 + 1))
    
    quy_nam = "Q" + pd.Series(quy).astype(str) + "/" + pd.Series(nam).astype(str)
    quy_nam = np.where((quy == 0) | (nam == 0), "", quy_nam)
    
    phi_goc = pd.to_numeric(df["Phi_bao_hiem_goc"].astype(str).str.replace(",", "", regex=False), errors='coerce').fillna(0.0).values
    phi_tai = pd.to_numeric(df["Phi_bao_hiem_tai"].astype(str).str.replace(",", "", regex=False), errors='coerce').fillna(0.0).values
    phi_giulai = phi_goc - phi_tai
    
    hieu_luc = df["hieu_luc"].astype(str).str.strip().str.lower().values
    
    for q in four_last_quarters:
        mask_q = (quy_nam == q)
        
        phi_goc_sum = phi_goc[mask_q].sum()
        phi_giulai_sum = phi_giulai[mask_q].sum()
        
        mask_giam = mask_q & (hieu_luc == "het")
        giam_goc_sum = phi_goc[mask_giam].sum()
        giam_giulai_sum = phi_giulai[mask_giam].sum()
        giam_tai_sum = giam_goc_sum - giam_giulai_sum
        
        summary_rows.append({
            "Quy": q,
            "Phi_bao_hiem_goc": phi_goc_sum,
            "Phi_bao_hiem_giu_lai": phi_giulai_sum,
            "Giam_phi_bao_hiem_goc": giam_goc_sum,
            "Giam_phi_bao_hiem_giu_lai": giam_giulai_sum,
            "Giam_phi_bao_hiem_tai": giam_tai_sum
        })
        
    return pd.DataFrame(summary_rows)

def calculate_tttbvv_summary_df(df: pd.DataFrame, ky_available: list[int], ty_gia: pd.DataFrame, dpnv_date: datetime.date, four_last_quarters: list[str]) -> pd.DataFrame:
    summary_rows = []
    
    ty_gia_dict_usd = dict(zip(ty_gia["Thoi_gian"].astype(str), ty_gia["USD"]))
    ty_gia_dict_eur = dict(zip(ty_gia["Thoi_gian"].astype(str), ty_gia["EUR"]))
    fallback_usd = ty_gia["USD"].iloc[-1] if not ty_gia.empty else 1.0
    fallback_eur = ty_gia["EUR"].iloc[-1] if not ty_gia.empty else 1.0
    
    dpnv = pd.to_datetime(dpnv_date)
    
    ret_col = "Ty_le_giu_lai_cua_BHBV" if "Ty_le_giu_lai_cua_BHBV" in df.columns else \
              ("Ty_le_giu_lai_cua_BHBV_checked" if "Ty_le_giu_lai_cua_BHBV_checked" in df.columns else None)
              
    for ky in ky_available:
        sh = f"Ky_phi{ky}"
        
        thang = pd.to_numeric(df[f"Ky_phi_{ky}_Ghi_Thang"], errors='coerce').fillna(0).astype(int)
        nam = pd.to_numeric(df[f"Ky_phi_{ky}_Ghi_Nam"], errors='coerce').fillna(0).astype(int)
        
        quy = np.where(thang == 0, 0, (thang - 1) // 3 + 1)
        quy_nam = "Q" + pd.Series(quy).astype(str) + "/" + pd.Series(nam).astype(str)
        quy_nam = np.where((quy == 0) | (nam == 0), "", quy_nam)
        
        rate_usd = pd.Series(quy_nam).map(ty_gia_dict_usd).fillna(fallback_usd).fillna(0.0).values
        rate_eur = pd.Series(quy_nam).map(ty_gia_dict_eur).fillna(fallback_eur).fillna(0.0).values
        
        vnd = pd.to_numeric(df[f"Ky_phi_{ky}_So_tien_VND"].astype(str).str.replace(",", "", regex=False), errors='coerce').fillna(0.0).values
        usd = pd.to_numeric(df[f"Ky_phi_{ky}_So_tien_USD"].astype(str).str.replace(",", "", regex=False), errors='coerce').fillna(0.0).values
        eur = pd.to_numeric(df[f"Ky_phi_{ky}_So_tien_EUR"].astype(str).str.replace(",", "", regex=False), errors='coerce').fillna(0.0).values
        
        phi_goc = vnd + usd * rate_usd + eur * rate_eur
        
        if ret_col:
            hi = pd.to_numeric(df[ret_col].astype(str).str.replace(",", "", regex=False), errors='coerce').fillna(0.0).values
            tile = np.where((hi == 0.0) | df[ret_col].isna(), 1.0, np.where(hi > 1.0, hi / 100.0, hi))
        else:
            tile = np.ones(len(df))
            
        phi_giulai = phi_goc * tile
        
        def parse_date_series(d_col, m_col, y_col):
            d_val = pd.to_numeric(df[d_col], errors='coerce')
            m_val = pd.to_numeric(df[m_col], errors='coerce')
            y_val = pd.to_numeric(df[y_col], errors='coerce')
            d_str = y_val.fillna(2000).astype(int).astype(str).str.zfill(4) + '-' + \
                    m_val.fillna(1).astype(int).astype(str).str.zfill(2) + '-' + \
                    d_val.fillna(1).astype(int).astype(str).str.zfill(2)
            parsed = pd.to_datetime(d_str, format='%Y-%m-%d', errors='coerce')
            parsed = parsed.where(d_val.notna() & m_val.notna() & y_val.notna(), pd.NaT)
            return parsed
            
        tu_date = parse_date_series(f"Ky_phi_{ky}_Tu_Ngay", f"Ky_phi_{ky}_Tu_Thang", f"Ky_phi_{ky}_Tu_Nam")
        den_date = parse_date_series(f"Ky_phi_{ky}_Den_Ngay", f"Ky_phi_{ky}_Den_Thang", f"Ky_phi_{ky}_Den_Nam")
        
        tong_so_ngay = (den_date - tu_date).dt.days + 1
        tong_so_ngay = np.where(tong_so_ngay < 0, 0, tong_so_ngay)
        
        so_ngay_da_qua = (dpnv - tu_date).dt.days + 1
        
        so_ngay_con_lai = np.where(
            so_ngay_da_qua < 0,
            tong_so_ngay,
            np.where(so_ngay_da_qua > tong_so_ngay, 0, tong_so_ngay - so_ngay_da_qua)
        )
        
        ratio = np.where(tong_so_ngay == 0, 0.0, so_ngay_con_lai / tong_so_ngay)
        
        du_phong_goc = ratio * phi_goc
        du_phong_giulai = ratio * phi_giulai
        
        for q in four_last_quarters:
            mask_q = (quy_nam == q)
            
            phi_goc_sum = phi_goc[mask_q].sum()
            phi_giulai_sum = phi_giulai[mask_q].sum()
            dpgoc_sum = du_phong_goc[mask_q].sum()
            dpgiu_sum = du_phong_giulai[mask_q].sum()
            dptai_sum = dpgoc_sum - dpgiu_sum
            
            summary_rows.append({
                "Ky_phi": sh,
                "Quy": q,
                "Phi_bao_hiem_goc": phi_goc_sum,
                "Phi_bao_hiem_giu_lai": phi_giulai_sum,
                "Du_phong_bao_hiem_goc": dpgoc_sum,
                "Du_phong_bao_hiem_giu_lai": dpgiu_sum,
                "Du_phong_bao_hiem_tai": dptai_sum
            })
            
    return pd.DataFrame(summary_rows)

def calculate_lt_summary_df(df: pd.DataFrame, ky_available: list[int], ty_gia: pd.DataFrame, dpnv_date: datetime.date, four_last_quarters: list[str]) -> pd.DataFrame:
    summary_rows = []
    
    ty_gia_dict_usd = dict(zip(ty_gia["Thoi_gian"].astype(str), ty_gia["USD"]))
    ty_gia_dict_eur = dict(zip(ty_gia["Thoi_gian"].astype(str), ty_gia["EUR"]))
    fallback_usd = ty_gia["USD"].iloc[-1] if not ty_gia.empty else 1.0
    fallback_eur = ty_gia["EUR"].iloc[-1] if not ty_gia.empty else 1.0
    
    dpnv = pd.to_datetime(dpnv_date)
    dpnv_nam = dpnv_date.year
    dpnv_thang = dpnv_date.month
    
    ret_col = "Ty_le_giu_lai_cua_BHBV" if "Ty_le_giu_lai_cua_BHBV" in df.columns else \
              ("Ty_le_giu_lai_cua_BHBV_checked" if "Ty_le_giu_lai_cua_BHBV_checked" in df.columns else None)
              
    quarters_list = ["Số dùng để tính"] + four_last_quarters
    
    for ky in ky_available:
        sh = f"Ky_phi{ky}"
        
        thang = pd.to_numeric(df[f"Ky_phi_{ky}_Ghi_Thang"], errors='coerce').fillna(0).astype(int)
        nam = pd.to_numeric(df[f"Ky_phi_{ky}_Ghi_Nam"], errors='coerce').fillna(0).astype(int)
        
        quy = np.where(thang == 0, 0, (thang - 1) // 3 + 1)
        quy_nam = "Q" + pd.Series(quy).astype(str) + "/" + pd.Series(nam).astype(str)
        quy_nam = np.where((quy == 0) | (nam == 0), "", quy_nam)
        
        rate_usd = pd.Series(quy_nam).map(ty_gia_dict_usd).fillna(fallback_usd).fillna(0.0).values
        rate_eur = pd.Series(quy_nam).map(ty_gia_dict_eur).fillna(fallback_eur).fillna(0.0).values
        
        vnd = pd.to_numeric(df[f"Ky_phi_{ky}_So_tien_VND"].astype(str).str.replace(",", "", regex=False), errors='coerce').fillna(0.0).values
        usd = pd.to_numeric(df[f"Ky_phi_{ky}_So_tien_USD"].astype(str).str.replace(",", "", regex=False), errors='coerce').fillna(0.0).values
        eur = pd.to_numeric(df[f"Ky_phi_{ky}_So_tien_EUR"].astype(str).str.replace(",", "", regex=False), errors='coerce').fillna(0.0).values
        
        phi_sau_dong = vnd + usd * rate_usd + eur * rate_eur
        
        if ret_col:
            hi = pd.to_numeric(df[ret_col].astype(str).str.replace(",", "", regex=False), errors='coerce').fillna(0.0).values
            tile = np.where((hi == 0.0) | df[ret_col].isna(), 1.0, np.where(hi > 1.0, hi / 100.0, hi))
        else:
            tile = np.ones(len(df))
            
        phi_giulai = phi_sau_dong * tile
        phi_tai = phi_sau_dong - phi_giulai
        
        def parse_date_series(d_col, m_col, y_col):
            d_val = pd.to_numeric(df[d_col], errors='coerce')
            m_val = pd.to_numeric(df[m_col], errors='coerce')
            y_val = pd.to_numeric(df[y_col], errors='coerce')
            d_str = y_val.fillna(2000).astype(int).astype(str).str.zfill(4) + '-' + \
                    m_val.fillna(1).astype(int).astype(str).str.zfill(2) + '-' + \
                    d_val.fillna(1).astype(int).astype(str).str.zfill(2)
            parsed = pd.to_datetime(d_str, format='%Y-%m-%d', errors='coerce')
            parsed = parsed.where(d_val.notna() & m_val.notna() & y_val.notna(), pd.NaT)
            return parsed
            
        tu_date = parse_date_series("Thoi_han_bao_hiem_Tu_Ngay", "Thoi_han_bao_hiem_Tu_Thang", "Thoi_han_bao_hiem_Tu_Nam")
        den_date = parse_date_series("Thoi_han_bao_hiem_Den_Ngay", "Thoi_han_bao_hiem_Den_Thang", "Thoi_han_bao_hiem_Den_Nam")
        
        tu_date_kp = parse_date_series(f"Ky_phi_{ky}_Tu_Ngay", f"Ky_phi_{ky}_Tu_Thang", f"Ky_phi_{ky}_Tu_Nam")
        den_date_kp = parse_date_series(f"Ky_phi_{ky}_Den_Ngay", f"Ky_phi_{ky}_Den_Thang", f"Ky_phi_{ky}_Den_Nam")
        ghi_date = parse_date_series(f"Ky_phi_{ky}_Ghi_Ngay", f"Ky_phi_{ky}_Ghi_Thang", f"Ky_phi_{ky}_Ghi_Nam")
        
        diff_dpnv_den = (dpnv - den_date_kp).dt.days
        c1 = np.where(diff_dpnv_den.isna(), 0, np.where(diff_dpnv_den >= 0, 0, 1))
        
        diff_den_tu = (den_date - tu_date).dt.days
        c2 = np.where(diff_den_tu.isna() | (diff_den_tu <= 365), 0, 1)
        
        c3 = np.where(df[f"Ky_phi_{ky}_So_tien_VND"].isna() & df[f"Ky_phi_{ky}_So_tien_USD"].isna() & df[f"Ky_phi_{ky}_So_tien_EUR"].isna(), 0, 1)
        
        diff_kp = (den_date_kp - tu_date_kp).dt.days
        c4 = np.where(diff_kp.isna(), 0, np.where(diff_kp > 0, 1, 0))
        
        c5 = np.where(df[f"Ky_phi_{ky}_Ghi_Thang"].isna() & df[f"Ky_phi_{ky}_Ghi_Nam"].isna(), 0, 1)
        
        diff_dpnv_tu = (dpnv - tu_date).dt.days
        c6 = np.where(diff_dpnv_tu.isna(), 0, np.where(diff_dpnv_tu >= 0, 1, 0))
        
        diff_ghi_dpnv = (ghi_date - dpnv).dt.days
        c7 = np.where(diff_ghi_dpnv.isna(), 0, np.where(diff_ghi_dpnv > 0, 0, 1))
        
        tonghop = np.where((c1 == 0) | (c2 == 0) | (c3 == 0) | (c4 == 0) | (c5 == 0) | (c6 == 0) | (c7 == 0), 0, 1)
        
        tu_nam_kp = pd.to_numeric(df[f"Ky_phi_{ky}_Tu_Nam"], errors='coerce').fillna(0).astype(int)
        tu_thang_kp = pd.to_numeric(df[f"Ky_phi_{ky}_Tu_Thang"], errors='coerce').fillna(0).astype(int)
        thu_tu = np.where(
            tonghop == 0,
            0,
            (dpnv_nam - tu_nam_kp) * 4 + (dpnv_thang - 1) // 3 + 1 - ((tu_thang_kp - 1) // 3 + 1) + 1
        )
        
        diff_kp_1 = (den_date_kp - tu_date_kp).dt.days + 1
        mau_so = np.where(diff_kp_1.isna(), 0.0, (diff_kp_1 / 365.0) * 8.0)
        
        huong_cu = np.where(tonghop == 0, mau_so, np.where(thu_tu <= 0, 0.0, thu_tu * 2.0 - 1.0))
        chua_huong = mau_so - huong_cu
        
        diff_den_dpnv = (den_date_kp - dpnv).dt.days
        chua_huong_dc = np.where(
            chua_huong >= 0.0,
            chua_huong,
            np.where(diff_den_dpnv.isna(), 0.0, (diff_den_dpnv / 365.0) * 8.0)
        )
        
        ts_final = chua_huong_dc
        ms_final = mau_so
        
        upr_ratio = np.where(ms_final == 0.0, 0.0, ts_final / ms_final)
        
        giu_chua_huong = phi_giulai * upr_ratio
        giu_duoc_huong = phi_giulai - giu_chua_huong
        
        tai_chua_huong = phi_tai * upr_ratio
        tai_duoc_huong = phi_tai - tai_chua_huong
        
        for q in quarters_list:
            if q == "Số dùng để tính":
                mask_q = (tonghop == 1)
            else:
                mask_q = (tonghop == 1) & (quy_nam == q)
                
            summary_rows.append({
                "Ky_phi": sh,
                "Quy": q,
                "Phi_bao_hiem_sau_dong": phi_sau_dong[mask_q].sum(),
                "Phi_bao_hiem_giu_lai": phi_giulai[mask_q].sum(),
                "Phi_tai_bao_hiem": phi_tai[mask_q].sum(),
                "Phi_bao_hiem_giu_lai_duoc_huong": giu_duoc_huong[mask_q].sum(),
                "Phi_tai_bao_hiem_duoc_huong": tai_duoc_huong[mask_q].sum(),
                "Phi_bao_hiem_giu_lai_chua_huong": giu_chua_huong[mask_q].sum(),
                "Phi_tai_bao_hiem_chua_huong": tai_chua_huong[mask_q].sum()
            })
            
    return pd.DataFrame(summary_rows)

def write_df_to_sheet(ws, df: pd.DataFrame, with_filter: bool = True):
    """Write pandas DataFrame to openpyxl worksheet."""
    # Write headers
    headers = list(df.columns)
    ws.append(headers)
    
    # Write rows
    for row in df.itertuples(index=False):
        row_vals = ["" if pd.isna(val) else val for val in row]
        ws.append(row_vals)
        
    if with_filter and df.shape[0] > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(df.shape[1])}{df.shape[0] + 1}"

def calculate_upr_for_file(quarter_id: str, file_name: str, group_code: str, dpnv_date: datetime.date) -> str:
    """
    Core UPR calculation implementation matching R's 4.1.ky_mau.R logic.
    Returns the path to the generated output Excel file.
    """
    parquet_path = os.path.join(DATA_ROOT, quarter_id, f"{group_code}.parquet")
    if not os.path.exists(parquet_path):
        dot_qid = quarter_id.replace("_", ".")
        alt_path = os.path.join(DATA_ROOT, dot_qid, f"{group_code}.parquet")
        if os.path.exists(alt_path):
            parquet_path = alt_path

    if not os.path.exists(parquet_path):
        raise FileNotFoundError(f"Merged parquet file not found at {parquet_path}")
        
    df = pd.read_parquet(parquet_path)
    
    # Load exchange rates
    ty_gia = load_ty_gia()
    
    # Determine type of UPR calculation
    is_vietjet = "vietjet" in file_name.lower()
    is_tttbvv = "tttbvv" in group_code.lower() or "tttbvv" in file_name.lower()
    is_lt = group_code.upper().endswith("_LT")
    
    # Create workbook in write-only mode to prevent massive memory usage (8M+ cells for PA_ST LOB)
    wb = openpyxl.Workbook(write_only=True)
        
    # Write exchange rates sheet
    ws_tygia = wb.create_sheet("Tygia")
    write_df_to_sheet(ws_tygia, ty_gia, with_filter=False)
    
    # Create Result sheet
    ws_result = wb.create_sheet("Result")
    
    # Get Ky_phi sheets available
    ky_available = []
    for col in df.columns:
        m = re.match(r"^Ky_phi_(\d+)_", col)
        if m:
            ky_available.append(int(m.group(1)))
    ky_available = sorted(list(set(ky_available)))
    
    dpnv_ngay = dpnv_date.day
    dpnv_thang = dpnv_date.month
    dpnv_nam = dpnv_date.year
    
    # Define VLOOKUP range & fallback cell
    num_rates = len(ty_gia)
    vlookup_range = f"Tygia!$A$1:$C${num_rates + 1}"
    fallback_cell = f"Tygia!$A${num_rates + 1}"
    
    # Extract four last quarters from exchange rates
    four_last_quarters = ty_gia.iloc[-4:, 0].tolist()
    
    # Output path - Named after group_code (LOB) instead of the raw uploaded file name
    quarter_out_dir = os.path.join(OUTPUT_EXCEL_ROOT, quarter_id)
    os.makedirs(quarter_out_dir, exist_ok=True)
    out_file_path = os.path.join(quarter_out_dir, f"{group_code}.xlsx")
    
    if is_vietjet:
        # --- Vietjet Branch ---
        columns_to_sum = [
            "Phi_bao_hiem_goc",
            "Phi_bao_hiem_giu_lai",
            "Giam_phi_bao_hiem_goc",
            "Giam_phi_bao_hiem_giu_lai",
            "Giam_phi_bao_hiem_tai"
        ]
        
        sheet_data = df.copy()
        extra_cols = ["Quy_phat_sinh_doanh_thu", "Quy_Nam", "Phi_bao_hiem_giu_lai"]
        
        ws_wj = wb.create_sheet("Vietjet")
        headers = list(sheet_data.columns) + extra_cols
        ws_wj.append(headers)
        
        def col_pos(col_name):
            return headers.index(col_name) + 1
            
        col_thang_let = int2col(col_pos("Thang_phat_sinh_doanh_thu"))
        col_quy_let = int2col(col_pos("Quy_phat_sinh_doanh_thu"))
        col_nam_let = int2col(col_pos("Nam_phat_sinh_doanh_thu"))
        goc_let = int2col(col_pos("Phi_bao_hiem_goc"))
        tai_let = int2col(col_pos("Phi_bao_hiem_tai"))
        
        n = len(sheet_data)
        col_lists = []
        for col in sheet_data.columns:
            lst = sheet_data[col].tolist()
            cleaned = ["" if (x is None or x is pd.NA or x is pd.NaT or x != x) else x for x in lst]
            col_lists.append(cleaned)
            
        for idx, row in enumerate(zip(*col_lists)):
            row_idx = idx + 2
            row_vals = list(row)
            
            formula_quy = f'=IF({col_thang_let}{row_idx}="",0,IF(VALUE({col_thang_let}{row_idx})=3,2, INT(({col_thang_let}{row_idx}-1)/3)+1))'
            formula_quynam = f'=CONCATENATE("Q",{col_quy_let}{row_idx},"/",{col_nam_let}{row_idx})'
            formula_giulai = f'={goc_let}{row_idx}-{tai_let}{row_idx}'
            
            row_vals.extend([formula_quy, formula_quynam, formula_giulai])
            ws_wj.append(row_vals)
            
        if n > 0:
            ws_wj.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{n + 1}"
            
        # Result Page Setup
        num_result_rows = len(four_last_quarters)
        
        # Write SUBTOTAL row
        subtotal_row = ["SUBTOTAL"]
        for j, col in enumerate(columns_to_sum):
            col_letter = int2col(j + 2)
            subtotal_row.append(f"=SUBTOTAL(9,{col_letter}3:{col_letter}{num_result_rows + 2})")
        ws_result.append(subtotal_row)
        
        # Write Headers
        ws_result.append(["Quy"] + columns_to_sum)
        
        # Write Rows
        for i, q in enumerate(four_last_quarters):
            row_idx = i + 3
            row_vals = [q]
            
            for j in range(2):
                col_name = columns_to_sum[j]
                c_range = f"Vietjet!{int2col(col_pos(col_name))}2:{int2col(col_pos(col_name))}{n+1}"
                q_range = f"Vietjet!{int2col(col_pos('Quy_Nam'))}2:{int2col(col_pos('Quy_Nam'))}{n+1}"
                row_vals.append(f'=SUMIFS({c_range},{q_range},"{q}")')
                
            goc_range = f"Vietjet!{int2col(col_pos('Phi_bao_hiem_goc'))}2:{int2col(col_pos('Phi_bao_hiem_goc'))}{n+1}"
            hieu_luc_range = f"Vietjet!{int2col(col_pos('hieu_luc'))}2:{int2col(col_pos('hieu_luc'))}{n+1}"
            q_range = f"Vietjet!{int2col(col_pos('Quy_Nam'))}2:{int2col(col_pos('Quy_Nam'))}{n+1}"
            row_vals.append(f'=SUMIFS({goc_range},{hieu_luc_range},"het",{q_range},"{q}")')
            
            giu_range = f"Vietjet!{int2col(col_pos('Phi_bao_hiem_giu_lai'))}2:{int2col(col_pos('Phi_bao_hiem_giu_lai'))}{n+1}"
            row_vals.append(f'=SUMIFS({giu_range},{hieu_luc_range},"het",{q_range},"{q}")')
            
            row_vals.append(f"=D{row_idx}-E{row_idx}")
            ws_result.append(row_vals)
            
    elif is_tttbvv:
        # --- TTTBVV Branch ---
        columns_to_sum = [
            "Phi_bao_hiem_goc",
            "Phi_bao_hiem_giu_lai",
            "Du_phong_bao_hiem_goc",
            "Du_phong_bao_hiem_giu_lai",
            "Du_phong_bao_hiem_tai"
        ]
        
        if len(ky_available) == 0:
            raise ValueError(f"No installments found for TTTBVV file {file_name}")
            
        sheet_names = [f"Ky_phi{k}" for k in ky_available]
        n = len(df)
        
        dummy_ky = ky_available[0]
        fee_cols = [f"Ky_phi_{dummy_ky}_So_tien_VND", f"Ky_phi_{dummy_ky}_So_tien_USD", f"Ky_phi_{dummy_ky}_So_tien_EUR",
                    f"Ky_phi_{dummy_ky}_Tu_Ngay", f"Ky_phi_{dummy_ky}_Tu_Thang", f"Ky_phi_{dummy_ky}_Tu_Nam",
                    f"Ky_phi_{dummy_ky}_Den_Ngay", f"Ky_phi_{dummy_ky}_Den_Thang", f"Ky_phi_{dummy_ky}_Den_Nam",
                    f"Ky_phi_{dummy_ky}_Ghi_Ngay", f"Ky_phi_{dummy_ky}_Ghi_Thang", f"Ky_phi_{dummy_ky}_Ghi_Nam"]
        fee_cols_exist = [c for c in fee_cols if c in df.columns]
        base_cols = list(df.columns[:min(24, len(df.columns))])
        
        extra_cols = [
            "Thoi_diem_tinh_DPNV_Ngay", "Thoi_diem_tinh_DPNV_Thang","Thoi_diem_tinh_DPNV_Nam",
            "Quy_ghi_doanh_thu", "Nam_ghi_doanh_thu", "Quy_Nam","Phi_bao_hiem_goc","Ty_le_giu_lai_BHBV",
            "Phi_bao_hiem_giu_lai","Tong_so_ngay","So_ngay_da_qua","So_ngay_con_lai","Du_phong_bao_hiem_goc","Du_phong_bao_hiem_giu_lai"
        ]
        
        def write_tttbvv_installment_sheet(ws_kp, ky):
            fee_cols_ky = [f"Ky_phi_{ky}_So_tien_VND", f"Ky_phi_{ky}_So_tien_USD", f"Ky_phi_{ky}_So_tien_EUR",
                           f"Ky_phi_{ky}_Tu_Ngay", f"Ky_phi_{ky}_Tu_Thang", f"Ky_phi_{ky}_Tu_Nam",
                           f"Ky_phi_{ky}_Den_Ngay", f"Ky_phi_{ky}_Den_Thang", f"Ky_phi_{ky}_Den_Nam",
                           f"Ky_phi_{ky}_Ghi_Ngay", f"Ky_phi_{ky}_Ghi_Thang", f"Ky_phi_{ky}_Ghi_Nam"]
            fee_cols_exist_ky = [c for c in fee_cols_ky if c in df.columns]
            sheet_data = df[base_cols + fee_cols_exist_ky].copy()
            
            headers = list(sheet_data.columns) + extra_cols
            ws_kp.append(headers)
            
            def col_pos(col_name):
                return headers.index(col_name) + 1
                
            col_thang_let = int2col(col_pos(f"Ky_phi_{ky}_Ghi_Thang"))
            col_nam_let = int2col(col_pos(f"Ky_phi_{ky}_Ghi_Nam"))
            col_quy_let = int2col(col_pos("Quy_ghi_doanh_thu"))
            col_nam_ghi_let = int2col(col_pos("Nam_ghi_doanh_thu"))
            
            col_vnd_let = int2col(col_pos(f"Ky_phi_{ky}_So_tien_VND"))
            col_usd_let = int2col(col_pos(f"Ky_phi_{ky}_So_tien_USD"))
            col_eur_let = int2col(col_pos(f"Ky_phi_{ky}_So_tien_EUR"))
            col_quynam_let = int2col(col_pos("Quy_Nam"))
            
            ret_col = "Ty_le_giu_lai_cua_BHBV" if "Ty_le_giu_lai_cua_BHBV" in sheet_data.columns else \
                      ("Ty_le_giu_lai_cua_BHBV_checked" if "Ty_le_giu_lai_cua_BHBV_checked" in sheet_data.columns else None)
            if not ret_col:
                raise ValueError("Could not find Ty_le_giu_lai_cua_BHBV column in TTTBVV sheet data")
            hi_let = int2col(col_pos(ret_col))
            
            phi_goc_col_let = int2col(col_pos("Phi_bao_hiem_goc"))
            tile_col_let = int2col(col_pos("Ty_le_giu_lai_BHBV"))
            
            tu_ngay_let = int2col(col_pos(f"Ky_phi_{ky}_Tu_Ngay"))
            tu_thang_let = int2col(col_pos(f"Ky_phi_{ky}_Tu_Thang"))
            tu_nam_let = int2col(col_pos(f"Ky_phi_{ky}_Tu_Nam"))
            
            den_ngay_let = int2col(col_pos(f"Ky_phi_{ky}_Den_Ngay"))
            den_thang_let = int2col(col_pos(f"Ky_phi_{ky}_Den_Thang"))
            den_nam_let = int2col(col_pos(f"Ky_phi_{ky}_Den_Nam"))
            
            dpnv_ngay_col_let = int2col(col_pos("Thoi_diem_tinh_DPNV_Ngay"))
            dpnv_thang_col_let = int2col(col_pos("Thoi_diem_tinh_DPNV_Thang"))
            dpnv_nam_col_let = int2col(col_pos("Thoi_diem_tinh_DPNV_Nam"))
            
            a_col_let = int2col(col_pos("Tong_so_ngay"))
            b_col_let = int2col(col_pos("So_ngay_da_qua"))
            b_col_con_let = int2col(col_pos("So_ngay_con_lai"))
            c_col_goc_let = int2col(col_pos("Phi_bao_hiem_goc"))
            c_col_giu_let = int2col(col_pos("Phi_bao_hiem_giu_lai"))
            
            col_lists = []
            for col in sheet_data.columns:
                lst = sheet_data[col].tolist()
                cleaned = ["" if (x is None or x is pd.NA or x is pd.NaT or x != x) else x for x in lst]
                col_lists.append(cleaned)
                
            for idx, row in enumerate(zip(*col_lists)):
                row_idx = idx + 2
                row_vals = list(row)
                
                formula_quy = f'=IF({col_thang_let}{row_idx}="",0,INT(({col_thang_let}{row_idx}-1)/3)+1)'
                formula_nam = f'=VALUE({col_nam_let}{row_idx})'
                formula_quynam = f'=CONCATENATE("Q",{col_quy_let}{row_idx},"/",{col_nam_ghi_let}{row_idx})'
                
                formula_goc = (
                    f'=VALUE({col_vnd_let}{row_idx}) + VALUE({col_usd_let}{row_idx}) * IFERROR(VALUE(VLOOKUP({col_quynam_let}{row_idx}, {vlookup_range}, 2, 0)), '
                    f'VALUE(VLOOKUP({fallback_cell}, {vlookup_range}, 2, 0))) + VALUE({col_eur_let}{row_idx}) * '
                    f'IFERROR(VALUE(VLOOKUP({col_quynam_let}{row_idx}, {vlookup_range}, 3, 0)), VALUE(VLOOKUP({fallback_cell}, {vlookup_range}, 3, 0)))'
                )
                
                formula_tile = f'=IF(OR({hi_let}{row_idx}="",{hi_let}{row_idx}=0), 1, IF(VALUE({hi_let}{row_idx}) > 1, VALUE({hi_let}{row_idx}) / 100, VALUE({hi_let}{row_idx})))'
                formula_giulai = f'=IF({phi_goc_col_let}{row_idx}="", 0, VALUE({phi_goc_col_let}{row_idx}) * {tile_col_let}{row_idx})'
                
                formula_tongngay = f'=MAX(0,DATE({den_nam_let}{row_idx},{den_thang_let}{row_idx},{den_ngay_let}{row_idx}) - DATE({tu_nam_let}{row_idx},{tu_thang_let}{row_idx},{tu_ngay_let}{row_idx})+1)'
                formula_ngaydaqua = f'=DATE({dpnv_nam_col_let}{row_idx},{dpnv_thang_col_let}{row_idx},{dpnv_ngay_col_let}{row_idx}) - DATE({tu_nam_let}{row_idx},{tu_thang_let}{row_idx},{tu_ngay_let}{row_idx})+1'
                formula_ngayconlai = f'=IF({b_col_let}{row_idx}<0,{a_col_let}{row_idx},IF({b_col_let}{row_idx}>{a_col_let}{row_idx},0,{a_col_let}{row_idx}-{b_col_let}{row_idx}))'
                
                formula_dpgoc = f'={b_col_con_let}{row_idx}/{a_col_let}{row_idx}*{c_col_goc_let}{row_idx}'
                formula_dpgiu = f'={b_col_con_let}{row_idx}/{a_col_let}{row_idx}*{c_col_giu_let}{row_idx}'
                
                row_vals.extend([
                    dpnv_ngay, dpnv_thang, dpnv_nam,
                    formula_quy, formula_nam, formula_quynam,
                    formula_goc, formula_tile, formula_giulai,
                    formula_tongngay, formula_ngaydaqua, formula_ngayconlai,
                    formula_dpgoc, formula_dpgiu
                ])
                ws_kp.append(row_vals)
                
            if len(sheet_data) > 0:
                ws_kp.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(sheet_data) + 1}"
                
            return {name: col_pos(name) for name in headers}
        
        col_pos_map = None
        for ky in ky_available:
            ws_kp = wb.create_sheet(f"Ky_phi{ky}")
            col_pos_map = write_tttbvv_installment_sheet(ws_kp, ky)
            
        num_result_rows = len(sheet_names) * len(four_last_quarters)
        
        subtotal_row = ["SUBTOTAL", ""]
        for j, col in enumerate(columns_to_sum):
            col_letter = int2col(j + 3)
            subtotal_row.append(f"=SUBTOTAL(9,{col_letter}3:{col_letter}{num_result_rows + 2})")
        ws_result.append(subtotal_row)
        
        ws_result.append(["Ky_phi", "Quy"] + columns_to_sum)
        
        row_idx = 3
        for sh in sheet_names:
            for q in four_last_quarters:
                row_vals = [sh, q]
                
                for j in range(4):
                    col_name = columns_to_sum[j]
                    c_range = f"{sh}!{int2col(col_pos_map[col_name])}2:{int2col(col_pos_map[col_name])}{n+1}"
                    q_range = f"{sh}!{int2col(col_pos_map['Quy_Nam'])}2:{int2col(col_pos_map['Quy_Nam'])}{n+1}"
                    row_vals.append(f'=SUMIFS({c_range},{q_range},"{q}")')
                    
                row_vals.append(f"=E{row_idx}-F{row_idx}")
                ws_result.append(row_vals)
                row_idx += 1
                
    elif is_lt:
        # --- Long Term (LT) Branch ---
        columns_to_sum = [
            "Phi_bao_hiem_sau_dong",
            "Phi_bao_hiem_giu_lai",
            "Phi_tai_bao_hiem",
            "Phi_bao_hiem_giu_lai_duoc_huong",
            "Phi_tai_bao_hiem_duoc_huong",
            "Phi_bao_hiem_giu_lai_chua_huong",
            "Phi_tai_bao_hiem_chua_huong"
        ]
        
        if len(ky_available) == 0:
            raise ValueError(f"No installments found for LT file {file_name}")
            
        sheet_names = [f"Ky_phi{k}" for k in ky_available]
        n = len(df)
        
        extra_cols = [
            "Quy_ghi_nhan_doanh_thu", "Quy_ghi_nhan_doanh_thu_2024", "Quy_ghi_nhan_doanh_thu_2025", 
            "Thoi_diem_ghi_nhan_doanh_thu", "Check_01", "Check_02", "Check_03", "Check_04", "Check_05", 
            "Check_06", "Check_07", "Tổng hợp các tiêu chí", 
            "Thoi_diem_tinh_DPNV_Ngay", "Thoi_diem_tinh_DPNV_Thang", "Thoi_diem_tinh_DPNV_Nam", 
            "Thu_tu_Quy_DPNV", "Mau_so", "Tu_so_huong_cu", "Tu_so_chua_huong", 
            "Tu_so_huong_sau_dieu_chinh", "Tu_so_chua_huong_dieu_chinh", 
            "TS_chua_huong_SĐC_final", "MS_SĐC_final", 
            "Phi_bao_hiem_sau_dong", "Phi_bao_hiem_giu_lai", "Phi_tai_bao_hiem", 
            "Phi_bao_hiem_giu_lai_duoc_huong", "Phi_tai_bao_hiem_duoc_huong", 
            "Phi_bao_hiem_giu_lai_chua_huong", "Phi_tai_bao_hiem_chua_huong", 
            "Check_Phi_bao_hiem_giu_lai_chua_huong", "Check_Phi_tai_bao_hiem_chua_huong"
        ]
        
        def write_lt_installment_sheet(ws_kp, ky):
            fee_cols_ky = [f"Ky_phi_{ky}_So_tien_VND", f"Ky_phi_{ky}_So_tien_USD", f"Ky_phi_{ky}_So_tien_EUR",
                           f"Ky_phi_{ky}_Tu_Ngay", f"Ky_phi_{ky}_Tu_Thang", f"Ky_phi_{ky}_Tu_Nam",
                           f"Ky_phi_{ky}_Den_Ngay", f"Ky_phi_{ky}_Den_Thang", f"Ky_phi_{ky}_Den_Nam",
                           f"Ky_phi_{ky}_Ghi_Ngay", f"Ky_phi_{ky}_Ghi_Thang", f"Ky_phi_{ky}_Ghi_Nam"]
            fee_cols_exist_ky = [c for c in fee_cols_ky if c in df.columns]
            base_cols = list(df.columns[:min(24, len(df.columns))])
            sheet_data = df[base_cols + fee_cols_exist_ky].copy()
            
            headers = list(sheet_data.columns) + extra_cols
            ws_kp.append(headers)
            
            def col_pos(col_name):
                return headers.index(col_name) + 1
                
            col_thang_let = int2col(col_pos(f"Ky_phi_{ky}_Ghi_Thang"))
            col_nam_let = int2col(col_pos(f"Ky_phi_{ky}_Ghi_Nam"))
            col_quy_let = int2col(col_pos("Quy_ghi_nhan_doanh_thu"))
            col_nam_ghi_let = int2col(col_pos(f"Ky_phi_{ky}_Ghi_Nam"))
            
            col_den_ngay_let = int2col(col_pos(f"Ky_phi_{ky}_Den_Ngay"))
            col_den_thang_let = int2col(col_pos(f"Ky_phi_{ky}_Den_Thang"))
            col_den_nam_let = int2col(col_pos(f"Ky_phi_{ky}_Den_Nam"))
            
            col_dpnv_ngay_let = int2col(col_pos("Thoi_diem_tinh_DPNV_Ngay"))
            col_dpnv_thang_let = int2col(col_pos("Thoi_diem_tinh_DPNV_Thang"))
            col_dpnv_nam_let = int2col(col_pos("Thoi_diem_tinh_DPNV_Nam"))
            
            tu_ngay_let = int2col(col_pos("Thoi_han_bao_hiem_Tu_Ngay"))
            tu_thang_let = int2col(col_pos("Thoi_han_bao_hiem_Tu_Thang"))
            tu_nam_let = int2col(col_pos("Thoi_han_bao_hiem_Tu_Nam"))
            
            den_ngay_let = int2col(col_pos("Thoi_han_bao_hiem_Den_Ngay"))
            den_thang_let = int2col(col_pos("Thoi_han_bao_hiem_Den_Thang"))
            den_nam_let = int2col(col_pos("Thoi_han_bao_hiem_Den_Nam"))
            
            col_vnd_let = int2col(col_pos(f"Ky_phi_{ky}_So_tien_VND"))
            col_usd_let = int2col(col_pos(f"Ky_phi_{ky}_So_tien_USD"))
            col_eur_let = int2col(col_pos(f"Ky_phi_{ky}_So_tien_EUR"))
            col_quynam_let = int2col(col_pos("Thoi_diem_ghi_nhan_doanh_thu"))
            
            tu_nam_kp_let = int2col(col_pos(f"Ky_phi_{ky}_Tu_Nam"))
            tu_thang_kp_let = int2col(col_pos(f"Ky_phi_{ky}_Tu_Thang"))
            tu_ngay_kp_let = int2col(col_pos(f"Ky_phi_{ky}_Tu_Ngay"))
            den_nam_kp_let = int2col(col_pos(f"Ky_phi_{ky}_Den_Nam"))
            den_thang_kp_let = int2col(col_pos(f"Ky_phi_{ky}_Den_Thang"))
            den_ngay_kp_let = int2col(col_pos(f"Ky_phi_{ky}_Den_Ngay"))
            
            col_ghi_thang_let = int2col(col_pos(f"Ky_phi_{ky}_Ghi_Thang"))
            col_ghi_nam_let = int2col(col_pos(f"Ky_phi_{ky}_Ghi_Nam"))
            
            col_tu_nam_let = int2col(col_pos("Thoi_han_bao_hiem_Tu_Nam"))
            col_tu_thang_let = int2col(col_pos("Thoi_han_bao_hiem_Tu_Thang"))
            col_tu_ngay_let = int2col(col_pos("Thoi_han_bao_hiem_Tu_Ngay"))
            
            col_ghi_ngay_let = int2col(col_pos(f"Ky_phi_{ky}_Ghi_Ngay"))
            
            c1_c_let = int2col(col_pos("Check_01"))
            c2_c_let = int2col(col_pos("Check_02"))
            c3_c_let = int2col(col_pos("Check_03"))
            c4_c_let = int2col(col_pos("Check_04"))
            c5_c_let = int2col(col_pos("Check_05"))
            c6_c_let = int2col(col_pos("Check_06"))
            c7_c_let = int2col(col_pos("Check_07"))
            col_tonghop_let = int2col(col_pos("Tổng hợp các tiêu chí"))
            
            col_thu_tu_let = int2col(col_pos("Thu_tu_Quy_DPNV"))
            col_mau_so_let = int2col(col_pos("Mau_so"))
            col_huong_cu_let = int2col(col_pos("Tu_so_huong_cu"))
            col_chua_huong_let = int2col(col_pos("Tu_so_chua_huong"))
            col_chua_huong_dieu_chinh_let = int2col(col_pos("Tu_so_chua_huong_dieu_chinh"))
            
            ret_col = "Ty_le_giu_lai_cua_BHBV" if "Ty_le_giu_lai_cua_BHBV" in sheet_data.columns else \
                      ("Ty_le_giu_lai_cua_BHBV_checked" if "Ty_le_giu_lai_cua_BHBV_checked" in sheet_data.columns else None)
            if not ret_col:
                raise ValueError("Could not find Ty_le_giu_lai_cua_BHBV column in LT sheet data")
            hi_let = int2col(col_pos(ret_col))
            
            phi_col_let = int2col(col_pos("Phi_bao_hiem_sau_dong"))
            phi_giu_lai_let = int2col(col_pos("Phi_bao_hiem_giu_lai"))
            
            ms_final_let = int2col(col_pos("MS_SĐC_final"))
            ts_final_let = int2col(col_pos("TS_chua_huong_SĐC_final"))
            
            phi_giu_chua_huong_let = int2col(col_pos("Phi_bao_hiem_giu_lai_chua_huong"))
            phi_tai_let = int2col(col_pos("Phi_tai_bao_hiem"))
            phi_tai_chua_huong_let = int2col(col_pos("Phi_tai_bao_hiem_chua_huong"))
            
            ts_huong_sdc_let = int2col(col_pos("Tu_so_huong_sau_dieu_chinh"))
            phi_giu_duoc_huong_let = int2col(col_pos("Phi_bao_hiem_giu_lai_duoc_huong"))
            phi_tai_duoc_huong_let = int2col(col_pos("Phi_tai_bao_hiem_duoc_huong"))
            
            col_lists = []
            for col in sheet_data.columns:
                lst = sheet_data[col].tolist()
                cleaned = ["" if (x is None or x is pd.NA or x is pd.NaT or x != x) else x for x in lst]
                col_lists.append(cleaned)
                
            for idx, row in enumerate(zip(*col_lists)):
                row_idx = idx + 2
                row_vals = list(row)
                
                formula_quy = f'=IF({col_thang_let}{row_idx}="",0,INT(({col_thang_let}{row_idx}-1)/3)+1)'
                formula_quy_2024 = f'=IF(VALUE({col_nam_let}{row_idx})=2024,{col_quy_let}{row_idx},0)'
                formula_quy_2025 = f'=IF(VALUE({col_nam_let}{row_idx})=2025,{col_quy_let}{row_idx},0)'
                formula_quynam = f'=CONCATENATE("Q",{col_quy_let}{row_idx},"/",{col_nam_ghi_let}{row_idx})'
                
                formula_c1 = f'=IFERROR(IF(DATE({col_dpnv_nam_let}{row_idx},{col_dpnv_thang_let}{row_idx},{col_dpnv_ngay_let}{row_idx}) - DATE({col_den_nam_let}{row_idx},{col_den_thang_let}{row_idx},{col_den_ngay_let}{row_idx}) >= 0, 0, 1), 0)'
                formula_c2 = f'=IF(OR({tu_ngay_let}{row_idx}="",{tu_thang_let}{row_idx}="",{tu_nam_let}{row_idx}="",{den_ngay_let}{row_idx}="",{den_thang_let}{row_idx}="",{den_nam_let}{row_idx}=""),0,IF(DATE({den_nam_let}{row_idx},{den_thang_let}{row_idx},{den_nam_let}{row_idx}) - DATE({tu_nam_let}{row_idx},{tu_thang_let}{row_idx},{tu_ngay_let}{row_idx}) <= 365,0,1))'
                formula_c3 = f'=IF(AND({col_vnd_let}{row_idx}="",{col_usd_let}{row_idx}="",{col_eur_let}{row_idx}=""),0,1)'
                formula_c4 = f'=IFERROR(IF(DATE({den_nam_kp_let}{row_idx},{den_thang_kp_let}{row_idx},{den_ngay_kp_let}{row_idx})-DATE({tu_nam_kp_let}{row_idx},{tu_thang_kp_let}{row_idx},{tu_ngay_kp_let}{row_idx})>0,1,0),0)'
                formula_c5 = f'=IF(AND({col_ghi_thang_let}{row_idx}="",{col_ghi_nam_let}{row_idx}=""), 0,1)'
                formula_c6 = f'=IFERROR(IF(DATE({col_dpnv_nam_let}{row_idx},{col_dpnv_thang_let}{row_idx},{col_dpnv_ngay_let}{row_idx}) - DATE({col_tu_nam_let}{row_idx},{col_tu_thang_let}{row_idx},{col_tu_ngay_let}{row_idx})>=0,1,0),0)'
                formula_c7 = f'=IFERROR(IF(DATE({col_ghi_nam_let}{row_idx},{col_ghi_thang_let}{row_idx},{col_ghi_ngay_let}{row_idx}) - DATE({col_dpnv_nam_let}{row_idx},{col_dpnv_thang_let}{row_idx},{col_dpnv_ngay_let}{row_idx}) > 0, 0, 1), 0)'
                
                formula_tonghop = f'=IF(OR({c1_c_let}{row_idx}=0,{c2_c_let}{row_idx}=0,{c3_c_let}{row_idx}=0,{c4_c_let}{row_idx}=0,{c5_c_let}{row_idx}=0,{c6_c_let}{row_idx}=0,{c7_c_let}{row_idx}=0), 0, 1)'
                formula_thutu = f'=IF({col_tonghop_let}{row_idx}=0, 0, IFERROR(({col_dpnv_nam_let}{row_idx} - {tu_nam_kp_let}{row_idx}) * 4 + INT(({col_dpnv_thang_let}{row_idx} - 1) / 3) + 1 - (INT(({tu_thang_kp_let}{row_idx} - 1) / 3) + 1 )+1, 0))'
                formula_mauso = f'=IFERROR(((DATE({den_nam_kp_let}{row_idx},{den_thang_kp_let}{row_idx},{den_ngay_kp_let}{row_idx}) - DATE({tu_nam_kp_let}{row_idx},{tu_thang_kp_let}{row_idx},{tu_ngay_kp_let}{row_idx}) + 1) / 365) * 8, 0)'
                
                formula_huongcu = f'=IF({col_tonghop_let}{row_idx}=0, {col_mau_so_let}{row_idx}, IF({col_thu_tu_let}{row_idx}<=0, 0, {col_thu_tu_let}{row_idx}*2-1))'
                formula_chuahuong = f'={col_mau_so_let}{row_idx} - {col_huong_cu_let}{row_idx}'
                formula_chuahuongdc = f'=IFERROR(IF({col_chua_huong_let}{row_idx}>=0, {col_chua_huong_let}{row_idx}, ((DATE({den_nam_kp_let}{row_idx},{den_thang_kp_let}{row_idx},{den_ngay_kp_let}{row_idx}) - DATE({col_dpnv_nam_let}{row_idx},{col_dpnv_thang_let}{row_idx},{col_dpnv_ngay_let}{row_idx})) / 365) * 8),0)'
                formula_huongsdc = f'={col_mau_so_let}{row_idx} - {col_chua_huong_dieu_chinh_let}{row_idx}'
                
                formula_tsfinal = f'={col_chua_huong_dieu_chinh_let}{row_idx}'
                formula_msfinal = f'={col_mau_so_let}{row_idx}'
                
                formula_phisau = (
                    f'=VALUE({col_vnd_let}{row_idx}) + VALUE({col_usd_let}{row_idx}) * IFERROR(VALUE(VLOOKUP({col_quynam_let}{row_idx}, {vlookup_range}, 2, 0)), '
                    f'VALUE(VLOOKUP({fallback_cell}, {vlookup_range}, 2, 0))) + VALUE({col_eur_let}{row_idx}) * '
                    f'IFERROR(VALUE(VLOOKUP({col_quynam_let}{row_idx}, {vlookup_range}, 3, 0)), VALUE(VLOOKUP({fallback_cell}, {vlookup_range}, 3, 0)))'
                )
                
                formula_phigiu = f'=IF({phi_col_let}{row_idx}="", 0, VALUE({phi_col_let}{row_idx}) * IF(OR({hi_let}{row_idx}="",{hi_let}{row_idx}=0), 1, IF(VALUE({hi_let}{row_idx}) > 1, VALUE({hi_let}{row_idx}) / 100, VALUE({hi_let}{row_idx}))))'
                formula_phitai = f'=IF({phi_col_let}{row_idx}="", 0, VALUE({phi_col_let}{row_idx}) - VALUE({phi_giu_lai_let}{row_idx}))'
                
                formula_giuchuahuong = f'=IF(VALUE({ms_final_let}{row_idx})=0, 0, VALUE({phi_giu_lai_let}{row_idx}) * VALUE({ts_final_let}{row_idx}) / VALUE({ms_final_let}{row_idx}))'
                formula_giuduochuong = f'=VALUE({phi_giu_lai_let}{row_idx}) - VALUE({phi_giu_chua_huong_let}{row_idx})'
                formula_taichuahuong = f'=IF(VALUE({ms_final_let}{row_idx})=0, 0, VALUE({phi_tai_let}{row_idx}) * VALUE({ts_final_let}{row_idx}) / VALUE({ms_final_let}{row_idx}))'
                formula_taiduochuong = f'=VALUE({phi_tai_let}{row_idx}) - VALUE({phi_tai_chua_huong_let}{row_idx})'
                
                formula_checkgiu = f'=IF(VALUE({ms_final_let}{row_idx}) = 0, 0, VALUE({phi_giu_lai_let}{row_idx}) * (VALUE({ts_huong_sdc_let}{row_idx}) / VALUE({ms_final_let}{row_idx})) - VALUE({phi_giu_duoc_huong_let}{row_idx}))'
                formula_checktai = f'=IF(VALUE({ms_final_let}{row_idx}) = 0, 0, VALUE({phi_tai_let}{row_idx}) * (VALUE({ts_huong_sdc_let}{row_idx}) / VALUE({ms_final_let}{row_idx})) - VALUE({phi_tai_duoc_huong_let}{row_idx}))'
                
                row_vals.extend([
                    formula_quy, formula_quy_2024, formula_quy_2025,
                    formula_quynam,
                    formula_c1, formula_c2, formula_c3, formula_c4, formula_c5, formula_c6, formula_c7,
                    formula_tonghop,
                    dpnv_ngay, dpnv_thang, dpnv_nam,
                    formula_thutu, formula_mauso, formula_huongcu, formula_chuahuong,
                    formula_huongsdc, formula_chuahuongdc,
                    formula_tsfinal, formula_msfinal,
                    formula_phisau, formula_phigiu, formula_phitai,
                    formula_giuduochuong, formula_taiduochuong, formula_giuchuahuong, formula_taichuahuong,
                    formula_checkgiu, formula_checktai
                ])
                ws_kp.append(row_vals)
                
            if len(sheet_data) > 0:
                ws_kp.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(sheet_data) + 1}"
                
            return {name: col_pos(name) for name in headers}
            
        col_pos_map = None
        for ky in ky_available:
            ws_kp = wb.create_sheet(f"Ky_phi{ky}")
            col_pos_map = write_lt_installment_sheet(ws_kp, ky)
            
        quarters_list = ["Số dùng để tính"] + four_last_quarters
        num_result_rows = len(sheet_names) * len(quarters_list)
        
        subtotal_row = ["SUBTOTAL", ""]
        for j, col in enumerate(columns_to_sum):
            col_letter = int2col(j + 3)
            subtotal_row.append(f"=SUBTOTAL(9,{col_letter}3:{col_letter}{num_result_rows + 2})")
        ws_result.append(subtotal_row)
        
        ws_result.append(["Ky_phi", "Quy"] + columns_to_sum)
        
        row_idx = 3
        for sh in sheet_names:
            for q in quarters_list:
                row_vals = [sh, q]
                
                for j, col in enumerate(columns_to_sum):
                    c_range = f"{sh}!{int2col(col_pos_map[col])}2:{int2col(col_pos_map[col])}{n+1}"
                    criteria_range = f"{sh}!{int2col(col_pos_map['Tổng hợp các tiêu chí'])}2:{int2col(col_pos_map['Tổng hợp các tiêu chí'])}{n+1}"
                    q_range = f"{sh}!{int2col(col_pos_map['Thoi_diem_ghi_nhan_doanh_thu'])}2:{int2col(col_pos_map['Thoi_diem_ghi_nhan_doanh_thu'])}{n+1}"
                    
                    if q != "Số dùng để tính":
                        row_vals.append(f'=SUMIFS({c_range},{criteria_range},"1",{q_range},"{q}")')
                    else:
                        row_vals.append(f'=SUMIFS({c_range},{criteria_range},"1")')
                ws_result.append(row_vals)
                row_idx += 1
                
    else:
        # --- Short Term (ST) Branch ---
        columns_to_sum = [
            "Phi_bao_hiem_goc",
            "Phi_bao_hiem_giu_lai",
            "Giam_phi_bao_hiem_goc",
            "Giam_phi_bao_hiem_giu_lai",
            "Giam_phi_bao_hiem_tai"
        ]
        
        if len(ky_available) == 0:
            raise ValueError(f"No installments found for ST file {file_name}")
            
        sheet_names = [f"Ky_phi{k}" for k in ky_available]
        n = len(df)
        
        extra_cols = [
            "Thoi_diem_tinh_DPNV_Ngay", "Thoi_diem_tinh_DPNV_Thang","Thoi_diem_tinh_DPNV_Nam",
            "Quy_ghi_doanh_thu", "Nam_ghi_doanh_thu", "Quy_Nam","Phi_bao_hiem_goc","Ty_le_giu_lai_BHBV",
            "Phi_bao_hiem_giu_lai","Dem_ngay","Het_hieu_luc"
        ]
        
        def write_st_installment_sheet(ws_kp, ky):
            fee_cols_ky = [f"Ky_phi_{ky}_So_tien_VND", f"Ky_phi_{ky}_So_tien_USD", f"Ky_phi_{ky}_So_tien_EUR",
                           f"Ky_phi_{ky}_Tu_Ngay", f"Ky_phi_{ky}_Tu_Thang", f"Ky_phi_{ky}_Tu_Nam",
                           f"Ky_phi_{ky}_Den_Ngay", f"Ky_phi_{ky}_Den_Thang", f"Ky_phi_{ky}_Den_Nam",
                           f"Ky_phi_{ky}_Ghi_Ngay", f"Ky_phi_{ky}_Ghi_Thang", f"Ky_phi_{ky}_Ghi_Nam"]
            fee_cols_exist_ky = [c for c in fee_cols_ky if c in df.columns]
            base_cols = list(df.columns[:min(24, len(df.columns))])
            sheet_data = df[base_cols + fee_cols_exist_ky].copy()
            
            headers = list(sheet_data.columns) + extra_cols
            ws_kp.append(headers)
            
            def col_pos(col_name):
                return headers.index(col_name) + 1
                
            col_thang_let = int2col(col_pos(f"Ky_phi_{ky}_Ghi_Thang"))
            col_nam_let = int2col(col_pos(f"Ky_phi_{ky}_Ghi_Nam"))
            col_quy_let = int2col(col_pos("Quy_ghi_doanh_thu"))
            col_nam_ghi_let = int2col(col_pos("Nam_ghi_doanh_thu"))
            
            col_vnd_let = int2col(col_pos(f"Ky_phi_{ky}_So_tien_VND"))
            col_usd_let = int2col(col_pos(f"Ky_phi_{ky}_So_tien_USD"))
            col_eur_let = int2col(col_pos(f"Ky_phi_{ky}_So_tien_EUR"))
            col_quynam_let = int2col(col_pos("Quy_Nam"))
            
            ret_col = "Ty_le_giu_lai_cua_BHBV" if "Ty_le_giu_lai_cua_BHBV" in sheet_data.columns else \
                      ("Ty_le_giu_lai_cua_BHBV_checked" if "Ty_le_giu_lai_cua_BHBV_checked" in sheet_data.columns else None)
            if not ret_col:
                raise ValueError("Could not find Ty_le_giu_lai_cua_BHBV column in ST sheet data")
            hi_let = int2col(col_pos(ret_col))
            
            phi_goc_col_let = int2col(col_pos("Phi_bao_hiem_goc"))
            tile_col_let = int2col(col_pos("Ty_le_giu_lai_BHBV"))
            
            tu_ngay_let = int2col(col_pos("Thoi_han_bao_hiem_Tu_Ngay"))
            tu_thang_let = int2col(col_pos("Thoi_han_bao_hiem_Tu_Thang"))
            tu_nam_let = int2col(col_pos("Thoi_han_bao_hiem_Tu_Nam"))
            
            den_ngay_let = int2col(col_pos("Thoi_han_bao_hiem_Den_Ngay"))
            den_thang_let = int2col(col_pos("Thoi_han_bao_hiem_Den_Thang"))
            den_nam_let = int2col(col_pos("Thoi_han_bao_hiem_Den_Nam"))
            
            dpnv_ngay_col_let = int2col(col_pos("Thoi_diem_tinh_DPNV_Ngay"))
            dpnv_thang_col_let = int2col(col_pos("Thoi_diem_tinh_DPNV_Thang"))
            dpnv_nam_col_let = int2col(col_pos("Thoi_diem_tinh_DPNV_Nam"))
            
            col_lists = []
            for col in sheet_data.columns:
                lst = sheet_data[col].tolist()
                cleaned = ["" if (x is None or x is pd.NA or x is pd.NaT or x != x) else x for x in lst]
                col_lists.append(cleaned)
                
            for idx, row in enumerate(zip(*col_lists)):
                row_idx = idx + 2
                row_vals = list(row)
                
                formula_quy = f'=IF({col_thang_let}{row_idx}="",0,INT(({col_thang_let}{row_idx}-1)/3)+1)'
                formula_nam = f'=VALUE({col_nam_let}{row_idx})'
                formula_quynam = f'=CONCATENATE("Q",{col_quy_let}{row_idx},"/",{col_nam_ghi_let}{row_idx})'
                
                formula_goc = (
                    f'=VALUE({col_vnd_let}{row_idx}) + VALUE({col_usd_let}{row_idx}) * IFERROR(VALUE(VLOOKUP({col_quynam_let}{row_idx}, {vlookup_range}, 2, 0)), '
                    f'VALUE(VLOOKUP({fallback_cell}, {vlookup_range}, 2, 0))) + VALUE({col_eur_let}{row_idx}) * '
                    f'IFERROR(VALUE(VLOOKUP({col_quynam_let}{row_idx}, {vlookup_range}, 3, 0)), VALUE(VLOOKUP({fallback_cell}, {vlookup_range}, 3, 0)))'
                )
                
                formula_tile = f'=IF(OR({hi_let}{row_idx}="",{hi_let}{row_idx}=0), 1, IF(VALUE({hi_let}{row_idx}) > 1, VALUE({hi_let}{row_idx}) / 100, VALUE({hi_let}{row_idx})))'
                formula_giulai = f'=IF({phi_goc_col_let}{row_idx}="", 0, VALUE({phi_goc_col_let}{row_idx}) * {tile_col_let}{row_idx})'
                
                formula_demngay = f'=IF(OR({tu_ngay_let}{row_idx}="",{tu_thang_let}{row_idx}="",{tu_nam_let}{row_idx}="",{den_ngay_let}{row_idx}="",{den_thang_let}{row_idx}="",{den_nam_let}{row_idx}=""),0,IF(DATE({den_nam_let}{row_idx},{den_thang_let}{row_idx},{den_ngay_let}{row_idx}) - DATE({tu_nam_let}{row_idx},{tu_thang_let}{row_idx},{tu_ngay_let}{row_idx}) < 365,1,0))'
                formula_hethieuluc = f'=IF(OR({dpnv_ngay_col_let}{row_idx}="",{dpnv_thang_col_let}{row_idx}="",{dpnv_nam_col_let}{row_idx}="",{den_ngay_let}{row_idx}="",{den_thang_let}{row_idx}="",{den_nam_let}{row_idx}=""),0,IF(DATE({dpnv_nam_col_let}{row_idx},{dpnv_thang_col_let}{row_idx},{dpnv_ngay_col_let}{row_idx}) - DATE({den_nam_let}{row_idx},{den_thang_let}{row_idx},{den_ngay_let}{row_idx}) >=0,1))'
                
                row_vals.extend([
                    dpnv_ngay, dpnv_thang, dpnv_nam,
                    formula_quy, formula_nam, formula_quynam,
                    formula_goc, formula_tile, formula_giulai,
                    formula_demngay, formula_hethieuluc
                ])
                ws_kp.append(row_vals)
                
            if len(sheet_data) > 0:
                ws_kp.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(sheet_data) + 1}"
                
            return {name: col_pos(name) for name in headers}
            
        col_pos_map = None
        for ky in ky_available:
            ws_kp = wb.create_sheet(f"Ky_phi{ky}")
            col_pos_map = write_st_installment_sheet(ws_kp, ky)
            
        num_result_rows = len(sheet_names) * len(four_last_quarters)
        
        subtotal_row = ["SUBTOTAL", ""]
        for j, col in enumerate(columns_to_sum):
            col_letter = int2col(j + 3)
            subtotal_row.append(f"=SUBTOTAL(9,{col_letter}3:{col_letter}{num_result_rows + 2})")
        ws_result.append(subtotal_row)
        
        ws_result.append(["Ky_phi", "Quy"] + columns_to_sum)
        
        row_idx = 3
        for sh in sheet_names:
            for q in four_last_quarters:
                row_vals = [sh, q]
                
                for j in range(2):
                    col_name = columns_to_sum[j]
                    c_range = f"{sh}!{int2col(col_pos_map[col_name])}2:{int2col(col_pos_map[col_name])}{n+1}"
                    q_range = f"{sh}!{int2col(col_pos_map['Quy_Nam'])}2:{int2col(col_pos_map['Quy_Nam'])}{n+1}"
                    row_vals.append(f'=SUMIFS({c_range},{q_range},"{q}")')
                    
                goc_range = f"{sh}!{int2col(col_pos_map['Phi_bao_hiem_goc'])}2:{int2col(col_pos_map['Phi_bao_hiem_goc'])}{n+1}"
                dem_ngay_range = f"{sh}!{int2col(col_pos_map['Dem_ngay'])}2:{int2col(col_pos_map['Dem_ngay'])}{n+1}"
                hieu_luc_range = f"{sh}!{int2col(col_pos_map['Het_hieu_luc'])}2:{int2col(col_pos_map['Het_hieu_luc'])}{n+1}"
                q_range = f"{sh}!{int2col(col_pos_map['Quy_Nam'])}2:{int2col(col_pos_map['Quy_Nam'])}{n+1}"
                row_vals.append(f'=SUMIFS({goc_range},{dem_ngay_range},"1",{hieu_luc_range},"1",{q_range},"{q}")')
                
                giu_range = f"{sh}!{int2col(col_pos_map['Phi_bao_hiem_giu_lai'])}2:{int2col(col_pos_map['Phi_bao_hiem_giu_lai'])}{n+1}"
                row_vals.append(f'=SUMIFS({giu_range},{dem_ngay_range},"1",{hieu_luc_range},"1",{q_range},"{q}")')
                
                row_vals.append(f"=E{row_idx}-F{row_idx}")
                ws_result.append(row_vals)
                row_idx += 1
                
    wb.save(out_file_path)
    
    # Calculate the summary DataFrame in Python, save to parquet and SQLite DB
    summary_path = out_file_path.replace(".xlsx", "_summary.parquet")
    try:
        if is_vietjet:
            df_summary = calculate_vietjet_summary_df(df, four_last_quarters)
        elif is_tttbvv:
            df_summary = calculate_tttbvv_summary_df(df, ky_available, ty_gia, dpnv_date, four_last_quarters)
        elif is_lt:
            df_summary = calculate_lt_summary_df(df, ky_available, ty_gia, dpnv_date, four_last_quarters)
        else:
            df_summary = calculate_st_summary_df(df, ky_available, ty_gia, dpnv_date, four_last_quarters)
        
        df_summary.to_parquet(summary_path)
        
        # Write to SQLite DB (Long_term, Short_term, or PA_TTTBVV)
        try:
            import sqlite3
            db_path = os.path.join(DATA_ROOT, f"{quarter_id.replace('_', '.')}.db")
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            if is_tttbvv:
                table_name = "PA_TTTBVV_Summary"
            elif is_lt:
                table_name = "Long_term"
            else:
                table_name = "Short_term"
                
            df_summary_db = df_summary.copy()
            df_summary_db.insert(0, "lob", group_code)
            
            # Drop empty columns
            df_summary_db = df_summary_db.dropna(how='all', axis=1)
            df_summary_db = df_summary_db.loc[:, [c for c in df_summary_db.columns if c is not None and str(c).strip() != ""]]
            
            # Delete existing rows for this LOB to avoid duplicates
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
            if cursor.fetchone():
                cursor.execute(f"DELETE FROM {table_name} WHERE lob = ?", (group_code,))
                conn.commit()
                
            df_summary_db.to_sql(table_name, conn, if_exists="append", index=False)
            conn.close()
        except Exception as db_err:
            print(f"Error writing summary to database for {file_name}: {db_err}")
            
    except Exception as e:
        print(f"Error pre-calculating summary for {file_name}: {e}")
        raise e
        
    recalculate_excel_file(out_file_path)
    return out_file_path

def calculate_upr_for_quarter(db: Session, quarter_id: str, file_ids: list[int] = None) -> dict:
    """
    Calculate UPR for a specific quarter.
    Processes selected file IDs or all files in status 'Merged'.
    """
    # Fetch calculations date
    param = db.query(models.AppParameter).filter(models.AppParameter.quarter_id == quarter_id).first()
    if param:
        dpnv_date = datetime.date(param.year, param.month, param.day)
    else:
        # Fallback date: end of quarter
        m = re.match(r"Q([1-4])_(\d{4})", quarter_id)
        if m:
            q, y = int(m.group(1)), int(m.group(2))
            if q == 1:
                dpnv_date = datetime.date(y, 3, 31)
            elif q == 2:
                dpnv_date = datetime.date(y, 6, 30)
            elif q == 3:
                dpnv_date = datetime.date(y, 9, 30)
            else:
                dpnv_date = datetime.date(y, 12, 31)
        else:
            dpnv_date = datetime.date.today()
            
    query = db.query(models.FileQueue).filter(models.FileQueue.quarter_id == quarter_id)
    if file_ids is not None:
        query = query.filter(models.FileQueue.id.in_(file_ids))
    else:
        query = query.filter(models.FileQueue.status == "Merged")
        
    files_to_calc = query.all()
    
    results = []
    for f in files_to_calc:
        f.status = "Calculating"
        db.commit()
        try:
            out_path = calculate_upr_for_file(
                quarter_id=f.quarter_id,
                file_name=f.file_name,
                group_code=f.group_code,
                dpnv_date=dpnv_date
            )
            f.status = "Calculated"
            db.commit()
            results.append({"file_id": f.id, "file_name": f.file_name, "status": "Success", "out_path": out_path})
        except Exception as e:
            f.status = "Error"
            db.commit()
            results.append({"file_id": f.id, "file_name": f.file_name, "status": "Error", "error": str(e)})
            
    return {"calculated": results}

def summarize_reports(db: Session, quarter_id: str, file_ids: list[int] = None) -> dict:
    """
    Consolidate calculated UPR from SQLite database tables into a summary workbook.
    """
    import sqlite3
    db_path = os.path.join(DATA_ROOT, f"{quarter_id.replace('_', '.')}.db")
    if not os.path.exists(db_path):
        raise ValueError(f"Database file for quarter {quarter_id} not found at {db_path}.")
        
    conn = sqlite3.connect(db_path)
    
    # Create final summarized workbook
    wb_new = openpyxl.Workbook()
    if "Sheet" in wb_new.sheetnames:
        wb_new.remove(wb_new["Sheet"])
        
    accounting_format = "#,##0"
    
    # We will read and write three groups: LongTerm, ShortTerm, PA_TTTBVV
    # Map sheet name to DB table name
    sheet_to_table = {
        "LongTerm": "Long_term",
        "ShortTerm": "Short_term",
        "PA_TTTBVV": "PA_TTTBVV_Summary"
    }
    
    for sheet_name, table_name in sheet_to_table.items():
        try:
            # Check if table exists in DB
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
            if not cursor.fetchone():
                print(f"Table {table_name} does not exist in DB, skipping.")
                continue
                
            df_sum = pd.read_sql_query(f"SELECT * FROM {table_name}", conn)
            if df_sum.empty:
                continue
                
            ws = wb_new.create_sheet(sheet_name)
            write_df_to_sheet(ws, df_sum, with_filter=True)
            
            # Apply format to numeric columns
            n_rows = len(df_sum)
            for r in range(2, n_rows + 2):
                for c in range(1, len(df_sum.columns) + 1):
                    val = ws.cell(row=r, column=c).value
                    if isinstance(val, (int, float)):
                        cell = ws.cell(row=r, column=c)
                        cell.number_format = accounting_format
                        
            # Auto-fit columns
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        except Exception as e:
            print(f"Error reading and writing table {table_name} for sheet {sheet_name}: {e}")
            continue
            
    conn.close()
    
    # Save output file with timestamp
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    summary_filename = f"Tong_Hop_Ketqua_{timestamp}.xlsx"
    summary_dir = os.path.join(OUTPUT_EXCEL_ROOT, quarter_id)
    os.makedirs(summary_dir, exist_ok=True)
    summary_path = os.path.join(summary_dir, summary_filename)
    
    # Check if we wrote at least one sheet
    if len(wb_new.sheetnames) == 0:
        wb_new.create_sheet("Empty")
        
    wb_new.save(summary_path)
    
    return {"file_name": summary_filename, "file_path": summary_path}
